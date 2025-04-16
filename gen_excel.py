# gen_excel.py

import os
import sys
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
import textwrap
import re

from utils import get_current_quote_yahoo

# Define Custom Fills
label_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")  # Light blue
data_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")   # Cornsilk

# Define custom fonts
title_font = Font(name="Times New Roman", size=14, bold=True, italic=True)
label_font = Font(name="Times New Roman", size=10, bold=True, italic=True)
data_tnr_font = Font(name = "Times New Roman", size=10)
data_tnr_italic_font = Font(name='Times New Roman', size=10, italic=True)
data_tnr_bold_font = Font(name = "Times New Roman", bold=True)
data_arial_font = Font(name = "Arial", size=10)
data_arial_bold_font = Font(name ="Arial", size=10, bold=True)
data_arial_italic_font = Font(name = "Arial", size=10, italic=True)

center_alignment = Alignment(horizontal="center", vertical="center")
right_alignment = Alignment(horizontal="right", vertical="center")
# Define a thin black border
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Define a thick black border for the outer edges
thick_border = Border(
    left=Side(style='thick', color='000000'),
    right=Side(style='thick', color='000000'),
    top=Side(style='thick', color='000000'),
    bottom=Side(style='thick', color='000000')
)

def apply_table_border(ws, row, start_col, end_col):
    """
    Applies a thin border around a group of cells in a specified row from start_col to end_col.

    :param ws: The worksheet object.
    :param row: The row number where the group is located.
    :param start_col: The starting column number of the group.
    :param end_col: The ending column number of the group.
    """
    # Define a thin black border
    thin_side = Side(style='thin', color='000000')
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        existing_border = cell.border
        
        # Initialize new border components with existing borders
        new_border = Border(
            top=existing_border.top,
            bottom=existing_border.bottom,
            left=existing_border.left,
            right=existing_border.right,
            diagonal=existing_border.diagonal,
            diagonal_direction=existing_border.diagonal_direction,
            outline=existing_border.outline,
            vertical=existing_border.vertical,
            horizontal=existing_border.horizontal
        )
        
        # Apply top and bottom borders to all cells in the group
        new_border.top = thin_side
        new_border.bottom = thin_side
        
        # Apply left border to the first cell and right border to the last cell
        if col == start_col:
            new_border.left = thin_side
        if col == end_col:
            new_border.right = thin_side
        
        # Assign the updated border back to the cell
        cell.border = new_border

def title_fill_range(ws, row_number, left_col, right_col):
        for cc in range(left_col, right_col + 1):
            cell = ws.cell(row_number, cc)
            if cell.fill.patternType is None:
                cell.fill = label_fill

def format_workbook(writer):
    """
    removes gridlines from all worksheets.
    """
    for sheetname in writer.book.sheetnames:
        ws = writer.book[sheetname]
        # Remove gridlines
        ws.sheet_view.showGridLines = False
        
def load_final_output(ticker):
    file_path = os.path.join("output", f"{ticker}_yoy_consolidated.json")
    if not os.path.exists(file_path):
        print(f"Error: {file_path} does not exist. Please generate {ticker}_yoy_consolidated.json first.")
        sys.exit(1)
    with open(file_path, "r") as f:
        data = json.load(f)
    return data

def create_xls(xls_filename):
    if os.path.exists(xls_filename):
        print(f"Overwriting existing file: {xls_filename}")
    return pd.ExcelWriter(xls_filename, engine='openpyxl', mode='w')

def write_summary_sheet(writer, final_output):
    wb = writer.book

    if 'Summary' not in wb.sheetnames:
        wb.create_sheet('Summary')
    ws = wb['Summary']

    summary_data = final_output["summary"]
    company_name = summary_data["company_name"]
    exchange = summary_data["exchange"]
    symbol = summary_data["symbol"]
    description = summary_data["description"]

    # Write and format the combined title in E1
    combined_title = f"{company_name.upper()} ({exchange}) - {symbol}"
    ws['E1'] = combined_title
    ws['E1'].font = title_font

    # Write "Company Description" in A4 with label formatting
    ws['A4'] = "Company Description"
    ws['A4'].font = label_font

    # Wrap and write the description
    wrapped_lines = textwrap.wrap(description, width=150)
    start_row = 5
    col = 2  # Column B
    for i, line in enumerate(wrapped_lines):
        cell = ws.cell(row=start_row + i, column=col, value=line)
        # Optionally, you can set font here if needed

def write_company_description(writer, final_output):
    reported_currency = final_output["summary"]["reported_currency"]
    cd_info = final_output["company_description"]
    cd_data = cd_info["data"]
    wb = writer.book

    if "Co. Desc" not in wb.sheetnames:
        wb.create_sheet("Co. Desc")
    ws = wb["Co. Desc"]
    ws.freeze_panes = "B1"

    # Write and format labels
    ws.cell(row=1, column=3, value="Currency").fill = label_fill
    ws.cell(row=1, column=3).font = label_font
    ws.cell(row=1, column=3).border = thin_border

    ws.cell(row=1, column=5, value="FY End").fill = label_fill
    ws.cell(row=1, column=5).font = label_font
    ws.cell(row=1, column=5).border = thin_border

    ws.cell(row=1, column=7, value="Stock Price").fill = label_fill
    ws.cell(row=1, column=7).font = label_font
    ws.cell(row=1, column=7).border = thin_border

    ws.cell(row=1, column=9, value="Market Cap").fill = label_fill
    ws.cell(row=1, column=9).font = label_font
    ws.cell(row=1, column=9).border = thin_border

    # Write and format data cells
    rc_cell = ws.cell(row=2, column=3, value=reported_currency)
    rc_cell.fill = data_fill
    rc_cell.font = label_font
    rc_cell.border = thin_border

    fiscal_year_end = cd_info.get("fiscal_year_end")
    fy_cell = ws.cell(row=2, column=5, value=fiscal_year_end)
    fy_cell.fill = data_fill
    fy_cell.font = label_font
    fy_cell.border = thin_border

    stock_price = to_float(cd_info.get("stock_price"))
    sp_cell = ws.cell(row=2, column=7, value=stock_price)
    sp_cell.fill = data_fill
    sp_cell.font = label_font
    sp_cell.number_format = '#,##0.00'
    sp_cell.border = thin_border

    market_cap = to_float(cd_info.get("marketCapitalization")) / 1_000_000
    mc_cell = ws.cell(row=2, column=9, value=market_cap)
    mc_cell.fill = data_fill
    mc_cell.font = label_font
    mc_cell.number_format = '#,##0'
    mc_cell.border = thin_border

    # Sort years for consistent column ordering
    sorted_years = sorted(cd_data.keys(), key=lambda x: int(x))

    # Determine the next two years
    if sorted_years:
        max_year = max(int(year) for year in sorted_years)
        new_years = [str(max_year + 1), str(max_year + 2)]
    else:
        new_years = ["2024", "2025"]

    # Append the new years to the sorted_years list
    all_years = sorted_years + new_years

    # Write the years at row=3 starting at column B
    start_col = 2  # Column B
    for i, year in enumerate(all_years):
        year_cell = ws.cell(row=3, column=start_col + i, value=year)
        year_cell.fill = label_fill
        year_cell.font = label_font
        year_cell.border = thin_border

    # Define metrics that should be displayed in millions
    million_scale_metrics = {
        "net_profit",
        "dividends_paid", 
        "shares_outstanding",
        "buyback",
        "share_equity",
        "long_term_debt"
    }

    # Define the metrics and their corresponding row positions
    metric_positions = {
        "net_profit": 4,            
        "diluted_eps": 5,           
        "operating_eps": 6,         
        "pe_ratio": 8,              
        "price_low": 9,             
        "price_high": 10,           
        "dividends_paid": 12,       
        "dividends_per_share": 13,  
        "avg_dividend_yield": 14,   
        "shares_outstanding": 16,   
        "buyback": 17,              
        "share_equity": 19,         
        "book_value_per_share": 20, 
        "long_term_debt": 22,       
        "roe": 24,                  
        "roc": 25                   
    }

    # Define human-readable labels for each metric
    metric_labels = {
        "net_profit": "Net Profit",
        "diluted_eps": "Diluted EPS",
        "operating_eps": "Operating EPS",
        "pe_ratio": "P/E Ratio",
        "price_low": "Yrly Price Low",
        "price_high": "Yrly Price High",
        "dividends_paid": "Dividends Paid",
        "dividends_per_share": "Dividends/Share",
        "avg_dividend_yield": "Avg Div Yield",
        "shares_outstanding": "Shares Outstanding",
        "buyback": "Buyback",
        "share_equity": "Share Equity",
        "book_value_per_share": "Book Value/Share",
        "long_term_debt": "Long-Term Debt",
        "roe": "ROE",
        "roc": "ROC"
    }

    # Updated number formats according to specifications
    number_formats = {
        "net_profit": '#,##0',          # Millions with commas, no $ no M
        "diluted_eps": '#,##0.00',     # Dollars and cents
        "operating_eps": '#,##0.00',    # Dollars and cents
        "pe_ratio": '#,##0.0',          # One decimal place
        "price_low": '#,##0.0',         # One decimal place
        "price_high": '#,##0.0',        # One decimal place
        "dividends_paid": '#,##0',      # Millions with commas, no $ no M
        "dividends_per_share": '#,##0.00',  # Dollars and cents
        "avg_dividend_yield": '0.00%',   # Percentage with two decimal places
        "shares_outstanding": '#,##0',   # Millions with commas, no $ no M
        "buyback": '#,##0',             # Millions with commas, no $ no M
        "share_equity": '#,##0',        # Millions with commas, no $ no M
        "book_value_per_share": '#,##0.00',  # Dollars and cents
        "long_term_debt": '#,##0',      # Millions with commas, no $ no M
        "roe": '0.0%',                  # Percentage with one decimal place
        "roc": '0.0%'                   # Percentage with one decimal place
    }

    # Write metric labels in column A with label formatting
    for metric, metric_row in metric_positions.items():
        label_cell = ws.cell(row=metric_row, column=1, value=metric_labels[metric])
        label_cell.fill = label_fill
        label_cell.font = label_font
        label_cell.border = thin_border

    # For each metric, write the data for each year in the specified rows
    for metric, metric_row in metric_positions.items():
        for i, year in enumerate(all_years):
            col = start_col + i
            col_letter = get_column_letter(col)
            data_cell = None

            # Determine if we should use formulas based on the metric
            use_formula = True
            
            if metric == "pe_ratio":
                formula = f"=(({col_letter}{metric_positions['price_low']}+{col_letter}{metric_positions['price_high']})/2)/{col_letter}{metric_positions['operating_eps']}"
            
            elif metric == "buyback" and i > 0:  # Skip first year as we need a previous year to compare
                prev_col = get_column_letter(col-1)
                formula = f"=({prev_col}{metric_positions['shares_outstanding']}-{col_letter}{metric_positions['shares_outstanding']})*({col_letter}{metric_positions['price_low']}+{col_letter}{metric_positions['price_high']})/2"
            elif metric == "buyback" and i == 0:  # For the first year, we can't calculate buyback
                formula = ""
                use_formula = False
                value = "N/A"
            
            elif metric == "dividends_per_share":
                formula = f"={col_letter}{metric_positions['dividends_paid']}/{col_letter}{metric_positions['shares_outstanding']}"
            
            elif metric == "avg_dividend_yield":
                formula = f"={col_letter}{metric_positions['dividends_per_share']}/((({col_letter}{metric_positions['price_low']}+{col_letter}{metric_positions['price_high']})/2))"
            
            elif metric == "book_value_per_share":
                formula = f"={col_letter}{metric_positions['share_equity']}/{col_letter}{metric_positions['shares_outstanding']}"
            
            elif metric == "roe":
                formula = f"={col_letter}{metric_positions['net_profit']}/{col_letter}{metric_positions['share_equity']}"
            
            elif metric == "roc":
                formula = f"={col_letter}{metric_positions['net_profit']}/({col_letter}{metric_positions['share_equity']}+{col_letter}{metric_positions['long_term_debt']})"
            
            else:
                use_formula = False
            
            # Apply either formula or value based on the decision
            if use_formula:
                data_cell = ws.cell(row=metric_row, column=col, value=formula)
                # For past years, use normal font; for future years, use italic font
                data_cell.font = data_tnr_italic_font if year in new_years else data_tnr_font
            else:
                # Get the value from data if no formula is used
                if not (metric == "buyback" and i == 0):  # Skip if it's buyback for the first year
                    value = cd_data.get(year, {}).get(metric)
                    if value is not None and metric in million_scale_metrics:
                        value = value / 1_000_000
                
                data_cell = ws.cell(row=metric_row, column=col, value=value)
                data_cell.font = data_tnr_italic_font if year in new_years else data_tnr_font
            
            # Apply common formatting to all cells
            data_cell.fill = data_fill
            data_cell.border = thin_border
            
            # Apply number format if defined
            if metric in number_formats:
                data_cell.number_format = number_formats[metric]
            
            # Apply right alignment for all data cells
            data_cell.alignment = right_alignment
            
def write_analyses_sheet(writer, final_output):
    reported_currency = final_output["summary"]["reported_currency"]
    analyses = final_output["analyses"]
    inv_char = analyses["investment_characteristics"]
    data = analyses["data"]
    wb = writer.book

    if "Analyses" not in wb.sheetnames:
        wb.create_sheet("Analyses")
    ws = wb["Analyses"]
    ws.freeze_panes = "B1"

    # Write and format the "Investment Characteristics" title
    ic_cell = ws.cell(row=1, column=6, value="Investment Characteristics (in mlns "+reported_currency + ")")
    ic_cell.fill = label_fill
    ic_cell.font = title_font
    apply_table_border(ws, 1, 5, 9)
    title_fill_range(ws, 1, 5, 9)

    # Extract required fields and write labels with formatting
    labels_with_positions = [
        (3, 3, "Earnings Analysis:"),
        (5, 4, "Growth Rate %:"),
        (7, 4, "Quality %:"),
        (3, 8, "Use Of Earnings Analysis:"),
        (5, 9, "Avg Div Payout Rate:"),
        (7, 9, "Avg Stk Buyback Rate:"),
        (13, 3, "Sales Analysis:"),
        (15, 4, "Growth Rate %:"),
        (17, 4, "Growth Rate PS %:"),
        (13, 8, "Sales Analysis (last 5 yrs.):"),
        (15, 9, "Growth Rate %:"),
        (17, 9, "Growth Rate PS %:")
    ]

    for row, col, text in labels_with_positions:
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = label_fill
        cell.font = label_font
        apply_table_border(ws, row, col, col+1)
        title_fill_range(ws, row, col, col+1)

    # Fetch data for the investment characteristics
    # These will be replaced with formulas
    growth_rate_operating_eps = inv_char["earnings_analysis"].get("growth_rate_percent_operating_eps")
    quality_percent = inv_char["earnings_analysis"].get("quality_percent")
    avg_div_payout = inv_char["use_of_earnings_analysis"].get("avg_dividend_payout_percent")
    avg_stk_buyback = inv_char["use_of_earnings_analysis"].get("avg_stock_buyback_percent")

    growth_rate_rev = inv_char["sales_analysis"].get("growth_rate_percent_revenues")
    growth_rate_sps = inv_char["sales_analysis"].get("growth_rate_percent_sales_per_share")
    growth_rate_rev_5y = inv_char["sales_analysis_last_5_years"].get("growth_rate_percent_revenues")
    growth_rate_sps_5y = inv_char["sales_analysis_last_5_years"].get("growth_rate_percent_sales_per_share")

    # Get sorted years for Co. Desc references
    sorted_years = sorted(data.keys(), key=lambda x: int(x))
    
    # Determine the next two years
    if sorted_years:
        max_year = max(int(year) for year in sorted_years)
        new_years = [str(max_year + 1), str(max_year + 2)]
    else:
        new_years = ["2024", "2025"]

    # Find the first and last year columns in Co. Desc sheet (for CAGR calculations)
    # Years in Co. Desc start at column B (2)
    first_year_col = 2  # Column B
    last_year_col = first_year_col + len(sorted_years) - 1  # Last historical year
    first_year_letter = get_column_letter(first_year_col)
    last_year_letter = get_column_letter(last_year_col)

    # Calculate the number of years for CAGR formula
    years_span = len(sorted_years) - 1 if len(sorted_years) > 1 else 1

    # Write data cells with formulas for specified metrics
    data_cells = {
        # Formula for Growth Rate % (Operating EPS): CAGR of operating_eps from Co. Desc
        # CAGR formula: (end_value/start_value)^(1/years) - 1
        (5, 6): f"=(('Co. Desc'!{last_year_letter}6/'Co. Desc'!{first_year_letter}6)^(1/{years_span})-1)",
        
        # Formula for Quality %: Avg diluted_eps / Avg operating_eps
        # Using AVERAGE function on the range of cells in Co. Desc
        (7, 6): f"=AVERAGE('Co. Desc'!{first_year_letter}5:{last_year_letter}5)/AVERAGE('Co. Desc'!{first_year_letter}6:{last_year_letter}6)",
        
        # Formula for Avg Div Payout Rate: Avg dividends_per_share / Avg operating_eps
        (5, 11): f"=AVERAGE('Co. Desc'!{first_year_letter}13:{last_year_letter}13)/AVERAGE('Co. Desc'!{first_year_letter}6:{last_year_letter}6)",
        
        # Keep the original value for Avg Stock Buyback Rate
        (7, 11): avg_stk_buyback,
        
        # Keep original values for these metrics
        (15, 6): growth_rate_rev,
        (17, 6): growth_rate_sps,
        (15, 11): growth_rate_rev_5y,
        (17, 11): growth_rate_sps_5y
    }

    # Write the investment characteristics data (data cells)
    for (row, col), value in data_cells.items():
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = data_fill
        cell.font = data_tnr_bold_font
        cell.border = thin_border
        # Apply percentage format for these values
        cell.number_format = '0.0%'

    # Handle the data by years
    # Append the new years to the sorted_years list
    all_years = sorted_years + new_years
    
    start_col = 2  # Column B

    # Write years at row 9 and row 19 (year labels)
    for i, year in enumerate(all_years):
        y9_cell = ws.cell(row=9, column=start_col + i, value=year)
        y9_cell.fill = label_fill
        y9_cell.font = label_font
        y9_cell.border = thin_border

        y19_cell = ws.cell(row=19, column=start_col + i, value=year)
        y19_cell.fill = label_fill
        y19_cell.font = label_font
        y19_cell.border = thin_border

    # Define the metrics and their row positions
    metric_rows_1 = {
        "revenues": 10,
        "sales_per_share": 11
    }

    metric_rows_2 = {
        "op_margin_percent": 20,
        "tax_rate": 21,
        "depreciation": 22,
        "depreciation_percent": 23
    }

    # Add labels in column A for these metrics with label formatting
    additional_labels = {
        10: "Revenues",
        11: "Sales/Share",
        20: "Operating Margin %",
        21: "Tax Rate %",
        22: "Depreciation",
        23: "Depreciation %"
    }

    # Define the formatting rules for each metric
    number_formats = {
        "revenues": '#,##0',  # Millions with commas
        "sales_per_share": '#,##0.00',  # Dollars and cents
        "op_margin_percent": '0.0%',  # Percentage with one decimal place
        "tax_rate": '0.0%',  # Percentage with one decimal place
        "depreciation": '#,##0',  # Millions with commas
        "depreciation_percent": '0.0%'  # Percentage with one decimal place
    }

    # Define metrics that should be displayed in millions
    million_scale_metrics = {
        "revenues",
        "depreciation"
    }

    for row, label in additional_labels.items():
        cell = ws.cell(row=row, column=1, value=label)
        cell.fill = label_fill
        cell.font = label_font
        cell.border = thin_border

    # Write first set of metrics (data cells)
    for metric, row_num in metric_rows_1.items():
        for i, year in enumerate(all_years):
            col = start_col + i
            col_letter = get_column_letter(col)
            
            if year in new_years:
                if metric == "sales_per_share":
                    # revenue this year / shares outstanding from Co. Desc sheet
                    formula = f"={col_letter}{metric_rows_1['revenues']}/('Co. Desc'!{col_letter}16)"
                    cell = ws.cell(row=row_num, column=col, value=formula)
                else:
                    val = data.get(year, {}).get(metric)
                    if val is not None and metric in million_scale_metrics:
                        val = val / 1_000_000
                    cell = ws.cell(row=row_num, column=col, value=val)
            else:
                val = data.get(year, {}).get(metric)
                if val is not None and metric in million_scale_metrics:
                    val = val / 1_000_000
                cell = ws.cell(row=row_num, column=col, value=val)
            
            cell.fill = data_fill
            cell.font = data_tnr_italic_font if year in new_years else data_tnr_font
            cell.border = thin_border
            cell.number_format = number_formats[metric]

    # Write second set of metrics (data cells)
    for metric, row_num in metric_rows_2.items():
        for i, year in enumerate(all_years):
            col = start_col + i
            col_letter = get_column_letter(col)
            
            if year in new_years:
                if metric == "depreciation_percent":
                    # MODIFIED: depreciation this year / net profit from Co. Desc sheet
                    formula = f"={col_letter}{metric_rows_2['depreciation']}/('Co. Desc'!{col_letter}4)"
                    cell = ws.cell(row=row_num, column=col, value=formula)
                else:
                    val = data.get(year, {}).get(metric)
                    if val is not None and metric in million_scale_metrics:
                        val = val / 1_000_000
                    cell = ws.cell(row=row_num, column=col, value=val)
            else:
                if metric == "depreciation_percent":
                    # MODIFIED: Ensure depreciation_percent uses formula for historical years too
                    formula = f"={col_letter}{metric_rows_2['depreciation']}/('Co. Desc'!{col_letter}4)"
                    cell = ws.cell(row=row_num, column=col, value=formula)
                else:
                    val = data.get(year, {}).get(metric)
                    if val is not None and metric in million_scale_metrics:
                        val = val / 1_000_000
                    cell = ws.cell(row=row_num, column=col, value=val)
            
            cell.fill = data_fill
            cell.font = data_tnr_italic_font if year in new_years else data_tnr_font
            cell.border = thin_border
            cell.number_format = number_formats[metric]

def write_profit_desc_sheet(writer, final_output, no_add_da=False):
    """
    Write the profit description sheet with updated handling for operating earnings breakdowns
    and calculations based on formulas rather than pre-calculated data
    
    Parameters:
    writer: ExcelWriter object
    final_output: Dictionary containing the full output data
    """
    reported_currency = final_output["summary"]["reported_currency"]
    pd_info = final_output["profit_description"]
    pchar = pd_info["profit_description_characteristics"]
    pdata = pd_info["data"]
    wb = writer.book

    # Create or access the "Profit.Desc." sheet
    if "Profit.Desc." not in wb.sheetnames:
        wb.create_sheet("Profit.Desc.")
    ws = wb["Profit.Desc."]
    ws.freeze_panes = "D1"

    # Write and format the title
    title_cell = ws.cell(row=1, column=4, value=f"Description & Analysis of Profitability (in mlns {reported_currency})")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_fill_range(ws, 1, 3, 10)
    apply_table_border(ws, 1, 3, 10)

    # Define the order of metrics to display
    metrics_order = [
        "revenues",
        "expenses",
        "ebitda",
        "amortization_depreciation",
        "free_cash_flow",
        "capex",
        "operating_earnings",
        "external_costs",
        "earnings",
        "non_gaap_earnings",  # Moved up to replace earnings_percent_revenue
        "dividend_paid",
        "dividend_paid_pct_fcf",
        "share_buybacks_from_stmt_cf",
        "net_biz_acquisition"
    ]

    # Define metric labels mapping
    metric_labels = {
        "revenues": "Net Revenues:",
        "expenses": "Internal Costs:",
        "ebitda": "EBITDA:",
        "free_cash_flow": "Free Cash Flow:",
        "operating_earnings": "Operating Margin:",
        "external_costs": "External Costs:",
        "dividend_paid": "Dividend Paid:",
        "dividend_paid_pct_fcf": "Dividend Paid % of FCF:",
        "share_buybacks_from_stmt_cf": "Share Buybacks (Stmt of CFs):",
        "net_biz_acquisition": "Net Biz Acquisitions:",
        "amortization_depreciation": "Amortization & Depreciation:",
        "capex": "Capital Expenditures:",
        "earnings": "Earnings:",
        "earnings_percent_revenue": "Earnings % of Revenue:",
        "non_gaap_earnings": "Non-GAAP Earnings:"  # Label remains the same
    }

    percent_metrics = ["dividend_paid_pct_fcf"]  # Removed earnings_percent_revenue since it's no longer used

    current_row = 5  # Starting row for metrics
    metric_rows = {}  # To track the row number for each metric
    breakdown_rows = {}  # To track the row numbers for each breakdown item

    # Step 1: Collect all unique breakdowns across all years
    all_revenue_breakdowns = set()
    all_expense_breakdowns = set()
    all_operating_earnings_breakdowns = set()
    all_external_costs_breakdowns = set()
    
    for year_data in pdata.values():
        # Handle revenue breakdowns
        revenues = year_data.get("revenues", {})
        if isinstance(revenues, dict) and "breakdown" in revenues:
            revenue_breakdown = revenues.get("breakdown", {})
            all_revenue_breakdowns.update(revenue_breakdown.keys())
            
        # Handle expense breakdowns
        expenses = year_data.get("expenses", {})
        if isinstance(expenses, dict) and "breakdown" in expenses:
            expense_breakdown = expenses.get("breakdown", {})
            all_expense_breakdowns.update(expense_breakdown.keys())
            
        # Handle operating earnings breakdowns
        operating_earnings = year_data.get("operating_earnings", {})
        if isinstance(operating_earnings, dict) and "breakdown" in operating_earnings:
            operating_breakdown = operating_earnings.get("breakdown", {})
            all_operating_earnings_breakdowns.update(operating_breakdown.keys())
            
        # Handle external costs breakdowns
        external_costs = year_data.get("external_costs", {})
        if isinstance(external_costs, dict) and "breakdown" in external_costs:
            external_costs_breakdown = external_costs.get("breakdown", {})
            all_external_costs_breakdowns.update(external_costs_breakdown.keys())

    # Step 2: Write the metrics and their labels
    for metric in metrics_order:
        # Get the label from the mapping
        label = metric_labels.get(metric, metric.capitalize() + ":")

        # Write the metric label cell with formatting
        label_cell = ws.cell(row=current_row, column=1, value=label)
        label_cell.fill = label_fill
        label_cell.font = label_font
        # Fill columns A to C for the label
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = label_fill
            cell.font = label_font
        metric_rows[metric] = current_row
        apply_table_border(ws, current_row, 1, 3)
        current_row += 1

        # Handle breakdowns based on metric type
        if metric == "revenues" and all_revenue_breakdowns:
            for bkey in sorted(all_revenue_breakdowns):
                breakdown_cell = ws.cell(row=current_row, column=2, value=bkey)
                breakdown_cell.font = Font(italic=True)
                breakdown_rows[(metric, bkey)] = current_row
                current_row += 1
        elif metric == "expenses" and all_expense_breakdowns:
            for bkey in sorted(all_expense_breakdowns):
                breakdown_cell = ws.cell(row=current_row, column=2, value=bkey)
                breakdown_cell.font = Font(italic=True)
                breakdown_rows[(metric, bkey)] = current_row
                current_row += 1
        elif metric == "operating_earnings" and all_operating_earnings_breakdowns:
            for bkey in sorted(all_operating_earnings_breakdowns):
                breakdown_cell = ws.cell(row=current_row, column=2, value=bkey)
                breakdown_cell.font = Font(italic=True)
                breakdown_rows[(metric, bkey)] = current_row
                current_row += 1
        elif metric == "external_costs" and all_external_costs_breakdowns:
            for bkey in sorted(all_external_costs_breakdowns):
                breakdown_cell = ws.cell(row=current_row, column=2, value=bkey)
                breakdown_cell.font = Font(italic=True)
                breakdown_rows[(metric, bkey)] = current_row
                current_row += 1

    # Step 3: Write Year Headers
    sorted_years = sorted(pdata.keys(), key=lambda x: int(x))
    start_col_for_years = 4  # Starting at column D

    for i, year in enumerate(sorted_years):
        year_col = start_col_for_years + i * 2  # Each year takes 2 columns
        y_cell = ws.cell(row=3, column=year_col, value=year)
        y_cell.fill = label_fill
        y_cell.font = label_font
        y_cell.border = thin_border
        y_cell.alignment = center_alignment

        # Add hyperlink if filing URL exists for this year
        filing_url = pdata.get(year, {}).get("filing_url")
        if filing_url:
            y_cell.hyperlink = filing_url
            y_cell.font = Font(name="Times New Roman", size=10, bold=True, italic=True, underline="single", color="0000FF")
            
    # Step 4: Write Metric Values and Breakdown Values using formulas where appropriate
    for i, year in enumerate(sorted_years):
        year_col = start_col_for_years + i * 2  # Column for the value
        year_data = pdata[year]
        
        # Map to analyses sheet column for this year
        analyses_col = 2 + i  # Starting at column B in Analyses sheet
        analyses_col_letter = get_column_letter(analyses_col)
        
        # Map to Co. Desc sheet column for this year
        co_desc_col = 2 + i  # Starting at column B in Co. Desc sheet
        co_desc_col_letter = get_column_letter(co_desc_col)
        
        for metric in metrics_order:
            metric_row = metric_rows.get(metric)
            
            # Skip if metric_row is None
            if metric_row is None:
                continue
                
            # Get the original metric data (for getting breakdowns later)
            original_metric_data = year_data.get(metric, {})
            
            # Handle specific metrics with formulas
            revenue_cell_ref = f"{get_column_letter(year_col)}{metric_rows['revenues']}"
            expenses_cell_ref = f"{get_column_letter(year_col)}{metric_rows['expenses']}"
            
            if metric == "amortization_depreciation":
                # Reference the depreciation row in Analyses sheet
                formula = f"='Analyses'!{analyses_col_letter}22"  # Assuming row 22 is depreciation in Analyses
                
                # Create and format the formula cell
                cell = ws.cell(row=metric_row, column=year_col, value=formula)
                cell.fill = data_fill
                cell.font = Font(name="Arial", italic=True)
                cell.border = thin_border
                cell.number_format = '#,##0'
            
            elif metric == "ebitda":
                # EBITDA calculation depends on no_add_da flag
                if no_add_da:
                    # Just Net Revenue - Internal Costs (no depreciation added back)
                    formula = f"={revenue_cell_ref}-{expenses_cell_ref}"
                else:
                    # Net Revenue - Internal Costs + Depreciation
                    amort_depr_cell_ref = f"{get_column_letter(year_col)}{metric_rows['amortization_depreciation']}"
                    formula = f"={revenue_cell_ref}-{expenses_cell_ref}+{amort_depr_cell_ref}"
                
                cell = ws.cell(row=metric_row, column=year_col, value=formula)
                cell.fill = data_fill
                cell.font = Font(name="Arial", italic=True)
                cell.border = thin_border
                cell.number_format = '#,##0'
            
            elif metric == "free_cash_flow":
                # Free Cash Flow = EBITDA - CapEx
                ebitda_cell_ref = f"{get_column_letter(year_col)}{metric_rows['ebitda']}"
                capex_cell_ref = f"{get_column_letter(year_col)}{metric_rows['capex']}"
                formula = f"={ebitda_cell_ref}+{capex_cell_ref}"
                
                cell = ws.cell(row=metric_row, column=year_col, value=formula)
                cell.fill = data_fill
                cell.font = Font(name="Arial", italic=True)
                cell.border = thin_border
                cell.number_format = '#,##0'
            
            elif metric == "operating_earnings":
                # Operating Margin
                if no_add_da:
                    amort_depr_cell_ref = f"{get_column_letter(year_col)}{metric_rows['amortization_depreciation']}"
                    formula = f"={revenue_cell_ref}-{expenses_cell_ref}-{amort_depr_cell_ref}"
                else:
                    # Net Revenue - Internal Costs 
                    formula = f"={revenue_cell_ref}-{expenses_cell_ref}"
                cell = ws.cell(row=metric_row, column=year_col, value=formula)
                cell.fill = data_fill
                cell.font = Font(name="Arial", italic=True)
                cell.border = thin_border
                cell.number_format = '#,##0'
                
            elif metric == "earnings":
                # Earnings = Operating Margin - External Costs
                op_margin_cell_ref = f"{get_column_letter(year_col)}{metric_rows['operating_earnings']}"
                ext_costs_cell_ref = f"{get_column_letter(year_col)}{metric_rows['external_costs']}"
                formula = f"={op_margin_cell_ref}-{ext_costs_cell_ref}"
                
                cell = ws.cell(row=metric_row, column=year_col, value=formula)
                cell.fill = data_fill
                cell.font = Font(name="Arial", italic=True)
                cell.border = thin_border
                cell.number_format = '#,##0'
                
            elif metric == "non_gaap_earnings":
                # Non-GAAP Earnings formula (retained from original)
                formula = f"='Co. Desc'!{co_desc_col_letter}6*'Co. Desc'!{co_desc_col_letter}16"
                
                cell = ws.cell(row=metric_row, column=year_col, value=formula)
                cell.fill = data_fill
                cell.font = Font(name="Arial", italic=True)
                cell.border = thin_border
                cell.number_format = '#,##0'
                
            elif metric == "share_buybacks_from_stmt_cf":
                # Reference the buyback line in Co. Desc sheet
                formula = f"='Co. Desc'!{co_desc_col_letter}17"
                
                cell = ws.cell(row=metric_row, column=year_col, value=formula)
                cell.fill = data_fill
                cell.font = Font(name="Arial", italic=True)
                cell.border = thin_border
                cell.number_format = '#,##0'
                
            elif metric == "dividend_paid_pct_fcf":
                # Dividend Paid % of FCF = Dividend Paid / Free Cash Flow
                dividend_cell_ref = f"{get_column_letter(year_col)}{metric_rows['dividend_paid']}"
                fcf_cell_ref = f"{get_column_letter(year_col)}{metric_rows['free_cash_flow']}"
                formula = f"={dividend_cell_ref}/{fcf_cell_ref}"
                
                cell = ws.cell(row=metric_row, column=year_col, value=formula)
                cell.fill = data_fill
                cell.font = Font(name="Arial", italic=True)
                cell.border = thin_border
                cell.number_format = '0.0%'
                
            else:
                # For other metrics, use the original data values
                metric_val = year_data.get(metric)
                if isinstance(metric_val, dict) and (
                    ("total_" + metric) in metric_val  # for operating_earnings
                    or ("total_" + metric.replace("_earnings", "")) in metric_val  # for revenues/external_costs
                ):
                    # Handle metrics with breakdowns (including operating earnings)
                    # Handle both operating_earnings and other metrics
                    if metric == "operating_earnings":
                        total_key = "total_operating_earnings"
                    else:
                        total_key = "total_" + metric.replace("_earnings", "")
                    
                    if total_key in metric_val:
                        val = to_float(metric_val[total_key])
                        if val is not None:
                            val = val / 1_000_000
                        cell = ws.cell(row=metric_row, column=year_col, value=val)
                        cell.fill = data_fill
                        cell.font = Font(name="Arial", italic=True)
                        cell.number_format = '#,##0'
                        cell.border = thin_border
                else:
                    # Handle metrics without breakdowns
                    val = to_float(metric_val)
                    if val is not None and metric not in percent_metrics:
                        val = val / 1_000_000
                    cell = ws.cell(row=metric_row, column=year_col, value=val)
                    cell.fill = data_fill
                    cell.font = Font(name="Arial", italic=True)
                    cell.border = thin_border
                    if metric not in percent_metrics:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '0.0%'
                
            # Write breakdown items - THIS IS THE KEY CHANGE: Always use original_metric_data for breakdowns
            # regardless of whether we used a formula for the main cell or not
            if isinstance(original_metric_data, dict) and "breakdown" in original_metric_data:
                breakdown_items = original_metric_data.get("breakdown", {})
                for (m, bkey), brow in breakdown_rows.items():
                    if m == metric and bkey in breakdown_items:
                        breakdown_val = breakdown_items[bkey]
                        if breakdown_val is not None:
                            breakdown_val = to_float(breakdown_val)
                            if breakdown_val is not None:
                                breakdown_val = breakdown_val / 1_000_000
                            bdata_cell = ws.cell(row=brow, column=year_col, value=breakdown_val)
                            bdata_cell.font = data_arial_italic_font
                            bdata_cell.number_format = '#,##0'
                            
                            # Add appropriate CAGR values
                            if metric == "revenues":
                                cagr_key = f"cagr_revenues_{bkey}_percent"
                                cagr_value = pchar.get("cagr_revenues_breakdown_percent", {}).get(cagr_key)
                                if cagr_value is not None:
                                    cagr_cell = ws.cell(row=brow, column=3, value=cagr_value)
                                    cagr_cell.font = Font(name="Arial", italic=True, size=8)
                                    cagr_cell.number_format = '0.0%'
                            elif metric == "operating_earnings":
                                cagr_key = f"cagr_operating_earnings_{bkey}_percent"
                                cagr_value = pchar.get("cagr_operating_earnings_breakdown_percent", {}).get(cagr_key)
                                if cagr_value is not None:
                                    cagr_cell = ws.cell(row=brow, column=3, value=cagr_value)
                                    cagr_cell.font = Font(name="Arial", italic=True, size=8)
                                    cagr_cell.number_format = '0.0%'

    # Step 5: Write CAGR Values
    cagr_map = {
        "revenues": "cagr_revenues_percent",
        "expenses": "cagr_total_expenses_percent",
        "ebitda": "cagr_ebitda_percent",
        "free_cash_flow": "cagr_free_cash_flow_percent",
        "operating_earnings": "cagr_operating_earnings_percent",
        "external_costs": "cagr_total_external_costs_percent",
        "earnings": "cagr_earnings_percent"
    }

    # Handle expense breakdown CAGRs
    expense_cagr_map = {
        "cost_of_revenue": "cagr_cost_of_revenue_percent",
        "research_and_development": "cagr_research_and_development_percent",
        "selling_marketing_general_admin": "cagr_selling_marketing_general_admin_percent"
    }

    # Write expense breakdown CAGRs
    for (metric, bkey), brow in breakdown_rows.items():
        if metric == "expenses" and bkey in expense_cagr_map:
            cagr_key = expense_cagr_map[bkey]
            cagr_value = pchar.get(cagr_key)
            if cagr_value is not None:
                cagr_cell = ws.cell(row=brow, column=3, value=cagr_value)
                cagr_cell.font = Font(name="Arial", italic=True, size=8)
                cagr_cell.number_format = '0.00%'
        elif metric == "external_costs":
            # Handle external costs breakdown CAGRs
            cagr_key = f"cagr_external_costs_{bkey}_percent"
            cagr_value = pchar.get("cagr_external_costs_breakdown_percent", {}).get(cagr_key)
            if cagr_value is not None:
                cagr_cell = ws.cell(row=brow, column=3, value=cagr_value)
                cagr_cell.font = Font(name="Arial", italic=True, size=8)
                cagr_cell.number_format = '0.00%'

    for metric, cagr_key in cagr_map.items():
        cagr_value = pchar.get(cagr_key)
        if cagr_value is not None and metric in metric_rows:
            row = metric_rows[metric]
            cagr_cell = ws.cell(row=row, column=3, value=cagr_value)
            cagr_cell.font = Font(name="Arial", italic=True, size=8)
            cagr_cell.number_format = '0.00%'

    # Step 6: Compute and Write Percentages for every metric related to revenues
    revenues_row = metric_rows.get("revenues")
    
    # Define which metrics should get percentage calculations
    percentage_metrics = [
        "expenses", "ebitda", "amortization_depreciation", "operating_earnings", 
        "external_costs", "earnings", "non_gaap_earnings"
    ]
    
    for i, year in enumerate(sorted_years):
        year_col = start_col_for_years + i * 2
        rev_cell_ref = f"{get_column_letter(year_col)}{revenues_row}"

        # Calculate percentages for all revenue breakdowns
        for bkey in all_revenue_breakdowns:
            br_key = ("revenues", bkey)
            if br_key in breakdown_rows:
                brow = breakdown_rows[br_key]
                metric_cell_ref = f"{get_column_letter(year_col)}{brow}"
                # Only display percentage if metric value is nonzero
                percent_formula = f"=IF({metric_cell_ref}<>0,{metric_cell_ref}/{rev_cell_ref},\"\")"
                
                percent_cell = ws.cell(row=brow, column=year_col + 1, value=percent_formula)
                percent_cell.font = Font(name="Arial", italic=True, size=8)
                percent_cell.number_format = '0.0%'

        # Calculate percentages for expense breakdowns
        for bkey in all_expense_breakdowns:
            br_key = ("expenses", bkey)
            if br_key in breakdown_rows:
                brow = breakdown_rows[br_key]
                metric_cell_ref = f"{get_column_letter(year_col)}{brow}"
                percent_formula = f"=IF({metric_cell_ref}<>0,{metric_cell_ref}/{rev_cell_ref},\"\")"
                
                percent_cell = ws.cell(row=brow, column=year_col + 1, value=percent_formula)
                percent_cell.font = Font(name="Arial", italic=True, size=8)
                percent_cell.number_format = '0.0%'

        # Calculate percentages for external costs breakdowns
        for bkey in all_external_costs_breakdowns:
            br_key = ("external_costs", bkey)
            if br_key in breakdown_rows:
                brow = breakdown_rows[br_key]
                metric_cell_ref = f"{get_column_letter(year_col)}{brow}"
                percent_formula = f"=IF({metric_cell_ref}<>0,{metric_cell_ref}/{rev_cell_ref},\"\")"
                
                percent_cell = ws.cell(row=brow, column=year_col + 1, value=percent_formula)
                percent_cell.font = Font(name="Arial", italic=True, size=8)
                percent_cell.number_format = '0.0%'

        # Calculate percentages for operating earnings breakdowns
        for bkey in all_operating_earnings_breakdowns:
            br_key = ("operating_earnings", bkey)
            if br_key in breakdown_rows:
                brow = breakdown_rows[br_key]
                metric_cell_ref = f"{get_column_letter(year_col)}{brow}"
                percent_formula = f"=IF({metric_cell_ref}<>0,{metric_cell_ref}/{rev_cell_ref},\"\")"
                
                percent_cell = ws.cell(row=brow, column=year_col + 1, value=percent_formula)
                percent_cell.font = Font(name="Arial", italic=True, size=8)
                percent_cell.number_format = '0.0%'

        # Calculate percentages for main metrics (use formulas for all)
        for tm in percentage_metrics:
            if tm in metric_rows:
                tm_row = metric_rows[tm]
                metric_cell_ref = f"{get_column_letter(year_col)}{tm_row}"
                percent_formula = f"=IF({metric_cell_ref}<>0,{metric_cell_ref}/{rev_cell_ref},\"\")"
                
                percent_cell = ws.cell(row=tm_row, column=year_col + 1, value=percent_formula)
                percent_cell.font = Font(name="Arial", italic=True, size=8)
                percent_cell.number_format = '0.0%'

def write_balance_sheet_sheet(writer, final_output):
    reported_currency = final_output["summary"]["reported_currency"]
    bs_info = final_output["balance_sheet"]
    bs_char = bs_info["balance_sheet_characteristics"]
    bs_data = bs_info["data"]
    wb = writer.book

    if "Balance Sht." not in wb.sheetnames:
        wb.create_sheet("Balance Sht.")
    ws = wb["Balance Sht."]
    ws.freeze_panes = "F1"
    # Write and format the title
    title_cell = ws.cell(row=1, column=4, value=f"Balance Sheet (in mlns {reported_currency}):")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_fill_range(ws, 1, 4,7)
    apply_table_border(ws, 1, 4, 7)

    # Extract characteristics
    cagr_assets = bs_char.get("cagr_total_assets_percent")
    cagr_liabilities = bs_char.get("cagr_total_liabilities_percent")
    cagr_equity = bs_char.get("cagr_total_shareholders_equity_percent")

    top_sections = [
        ("assets", "Assets:"),
        ("liabilities", "Liabilities:"),
        ("shareholders_equity", "Shareholder's Equity:")
    ]

    sorted_years = sorted(bs_data.keys(), key=lambda x: int(x))
    start_col_for_years = 6  # Column F

    # Write and format years in row 3
    for i, year in enumerate(sorted_years):
        year_col = start_col_for_years + i
        y_cell = ws.cell(row=3, column=year_col, value=year)
        y_cell.fill = label_fill
        y_cell.font = label_font
        y_cell.border = thin_border

    current_row = 5
    section_rows = {}

    for section_key, section_label in top_sections:
        # Main heading label in column A with label formatting
        sec_label_cell = ws.cell(row=current_row, column=1, value=section_label)
        sec_label_cell.fill = label_fill
        sec_label_cell.font = label_font
        title_fill_range(ws, current_row, 1,5)
        apply_table_border(ws, current_row, 1, 5)
        section_rows[section_key] = current_row

        first_year = sorted_years[0]
        section_data = bs_data[first_year].get(section_key, {})
        total_key = "total_" + section_key

        # Write total values for each year (data fill)
        for i, year in enumerate(sorted_years):
            year_col = start_col_for_years + i
            val = to_float(bs_data[year][section_key].get(total_key))
            if val is not None:
                val = val / 1_000_000
            d_cell = ws.cell(row=current_row, column=year_col, value=val)
            # This is the main heading's data row, so data fill it
            d_cell.fill = data_fill
            d_cell.font = data_arial_bold_font
            d_cell.border = thin_border
            # Apply number format
            if isinstance(val, (int, float)):
                d_cell.number_format = '#,##0'

        # Apply CAGR to column E
        if section_key == "assets":
            if cagr_assets is not None:
                ws.cell(row=section_rows["assets"], column=5, value=cagr_assets).number_format = '0.0%'
        elif section_key == "liabilities":
            if cagr_liabilities is not None:
                ws.cell(row=section_rows["liabilities"], column=5, value=cagr_liabilities).number_format = '0.0%'
        elif section_key == "shareholders_equity":
            if cagr_equity is not None:
                ws.cell(row=section_rows["shareholders_equity"], column=5, value=cagr_equity).number_format = '0.0%'

        current_row += 1

        # Breakdown items (no fill, no font)
        breakdown_data = section_data.get("breakdown", {})
        breakdown_rows = {}
        for bkey in breakdown_data.keys():
            # Breakdown label no fill
            ws.cell(row=current_row, column=2, value=bkey)
            breakdown_rows[bkey] = current_row
            current_row += 1

        # Write breakdown data (no fill)
        for bkey, brow in breakdown_rows.items():
            for i, year in enumerate(sorted_years):
                year_col = start_col_for_years + i
                val = to_float(bs_data[year][section_key]["breakdown"].get(bkey))
                if val is not None:
                    val = val / 1_000_000
                bdata_cell = ws.cell(row=brow, column=year_col, value=val)
                bdata_cell.font = data_arial_italic_font
                # Apply number format
                if isinstance(val, (int, float)):
                    bdata_cell.number_format = '#,##0'

        # Add a blank line before the next section
        current_row += 1

def write_studies_sheet(writer, final_output):
    reported_currency = final_output["summary"]["reported_currency"]
    studies = final_output.get("studies", {})
    analysis = studies.get("analysis_of_debt_levels", {})

    tdc = analysis.get("total_debt_capital", {})
    ltd = analysis.get("long_term_debt", {})
    nip = analysis.get("net_income_payback", {})
    anip = analysis.get("addback_net_inc_payback", {})

    wb = writer.book
    if "Studies" not in wb.sheetnames:
        wb.create_sheet("Studies")
    ws = wb["Studies"]

    # Write and format the title
    title_cell = ws.cell(row=1, column=4, value=f"Description & Analysis of Debt Levels (in mlns {reported_currency}):")
    title_cell.font = title_font
    apply_table_border(ws, 1 ,3, 10)
    title_fill_range(ws, 1, 3,10)

    # Summary section labels (apply label_fill and label_font)
    summary_labels = ["Summary:", "Total Debt-Capital:", "Long Term Debt-Cap.:", "Net Income Payback:", "Addback Net Inc Payback:"]
    summary_rows = [3, 7, 15, 23, 35]
    for label, row in zip(summary_labels, summary_rows):
        cell = ws.cell(row=row, column=1, value=label)
        cell.fill = label_fill
        cell.font = label_font
        cell.border = thin_border

    # Summary Section Texts (no font changes)
    ws.cell(row=3, column=2, value="Debt is a four-letter word.  Debt causes the years of repayment of capital to equity shareholders to stretch").font = data_arial_font
    ws.cell(row=4, column=2, value="out into the more distant future.  Even worse, debt can cause the best business model to become the").font = data_arial_font
    ws.cell(row=5, column=2, value="property of bondholders in a rough economic environment.").font = data_arial_font

    # Total Debt-Capital Section Texts
    ws.cell(row=7, column=2, value="The measure of total debt to total capital is useful when book value is a good measure of a firm's worth.  This").font = data_arial_font
    ws.cell(row=8, column=2, value="is particularly true of traditional businesses where property, plant and equipment are important.  Further, it").font = data_arial_font
    ws.cell(row=9, column=2, value="helps to have this ratio in capital intensive businesses with cyclical earnings.").font = data_arial_font

    # Total Debt-Capital Data
    ws.cell(row=11, column=3, value="Total Debt:").font = label_font
    val = tdc.get("total_debt")
    if val is not None:
        val = val / 1_000_000
    data_cell = ws.cell(row=11, column=4, value=val)
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font
    ws.cell(row=11, column=5, value="Here, deferred income taxes have been excluded.").font = data_arial_font

    ws.cell(row=12, column=3, value="Total Capital:").font = label_font
    val = tdc.get("total_capital")
    if val is not None:
        val = val / 1_000_000
    data_cell = ws.cell(row=12, column=4, value=val)
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font
    ws.cell(row=12, column=5, value="Here, deferred income taxes have been excluded.").font = data_arial_font

    ws.cell(row=13, column=3, value="Ratio:").font = label_font
    data_cell = ws.cell(row=13, column=4, value=tdc.get("total_debt_ratio"))
    data_cell.number_format = '0.0%'
    data_cell.font = data_arial_font

    # Long Term Debt-Cap. Section Texts
    ws.cell(row=15, column=2, value="The measure of long term debt to total capital is useful when total debt is distorted by the high presence").font = data_arial_font
    ws.cell(row=16, column=2, value="of current assets being financed by current liabilities.  Again, the measure works best within a traditional").font = data_arial_font
    ws.cell(row=17, column=2, value="industry setting.  The ratio helps position the equity shareholders.").font = data_arial_font

    # Long Term Debt-Cap. Data
    ws.cell(row=19, column=3, value="L. T. Debt:").font = label_font
    val = ltd.get("lt_debt")
    if val is not None:
        val = val / 1_000_000
    data_cell = ws.cell(row=19, column=4, value=val)
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font
    ws.cell(row=19, column=5, value="Here, the current liabilities have been excluded.").font = data_arial_font

    ws.cell(row=20, column=3, value="L. T. Capital:").font = label_font
    val = ltd.get("lt_capital")
    if val is not None:
        val = val / 1_000_000
    data_cell = ws.cell(row=20, column=4, value=val)
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font
    ws.cell(row=20, column=5, value="Here, the current liabilities have been excluded.").font = data_arial_font

    ws.cell(row=21, column=3, value="Ratio:").font = label_font
    data_cell = ws.cell(row=21, column=4, value=ltd.get("lt_debt_ratio"))
    data_cell.number_format = '0.0%'
    data_cell.font = data_arial_font

    # Net Income Payback Section Texts
    ws.cell(row=23, column=2, value="The measure of how quickly total debt is repaid by net income is a conservative measure, as it includes").font = data_arial_font
    ws.cell(row=24, column=2, value="debt such as current liabilities, that are financed by current assets and excludes some sources of cash, such").font = data_arial_font
    ws.cell(row=25, column=2, value="as noncash amortization numbers.").font = data_arial_font

    # Net Income Payback Data
    ws.cell(row=27, column=3, value="Total Debt:").font = label_font
    val = nip.get("total_debt")
    if val is not None:
        val = val / 1_000_000
    data_cell = ws.cell(row=27, column=4, value=val)
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font

    ws.cell(row=28, column=3, value="Net Income:").font = label_font
    val = nip.get("net_income")
    if val is not None:
        val = val / 1_000_000
    data_cell = ws.cell(row=28, column=4, value=val)
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font

    ws.cell(row=29, column=3, value="Years Payback:").font = label_font
    data_cell = ws.cell(row=29, column=4, value=nip.get("years_payback_total_debt"))
    data_cell.number_format = '#,##0.0'
    data_cell.font = data_arial_font

    ws.cell(row=31, column=3, value="L.T. Debt:").font = label_font
    val = nip.get("lt_debt")
    if val is not None:
        val = val / 1_000_000
    data_cell = ws.cell(row=31, column=4, value=val)
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font

    ws.cell(row=32, column=3, value="Net Income:").font = label_font
    val = nip.get("net_income")
    if val is not None:
        val = val / 1_000_000
    data_cell = ws.cell(row=32, column=4, value=val)
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font

    ws.cell(row=33, column=3, value="Years Payback:").font = label_font
    data_cell = ws.cell(row=33, column=4, value=nip.get("years_payback_lt_debt"))
    data_cell.number_format = '#,##0.0'
    data_cell.font = data_arial_font

    # Addback Net Inc Payback Section Texts
    ws.cell(row=35, column=2, value="The measure of how quickly debt is repaid by addback net income is a good measure, as it starts with GAAP").font = data_arial_font
    ws.cell(row=36, column=2, value="net income and adds back expenses on an after-tax basis that are clearly discretionary, such as business").font = data_arial_font
    ws.cell(row=37, column=2, value="acquisitions to better analyze the strength of the repayment stream.").font = data_arial_font

    # Addback Net Inc Payback Data
    ws.cell(row=39, column=3, value="L.T. Debt:").font = label_font
    val = anip.get("lt_debt")
    if val is not None:
        val = val / 1_000_000
    data_cell = ws.cell(row=39, column=4, value=val)
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font

    ws.cell(row=40, column=3, value="Net Income:").font = label_font
    # Get the historical years (exclude future forecast years)
    cd_data = final_output["company_description"]["data"]
    sorted_years = sorted(cd_data.keys(), key=lambda x: int(x))
    
    # Determine which years are historical vs forecast
    # In Co. Desc., years start at column B (2)
    start_col = 2  # Column B
    if sorted_years:
        # Get the last historical year
        # We want the last year with actual data, not forecast years
        last_year = sorted_years[-1]
        last_year_col = start_col + sorted_years.index(last_year)
        
        # Create a formula that multiplies operating EPS (row 6) by shares outstanding (row 16)
        # for the last historical year to get the net income
        formula = f"='Co. Desc'!{get_column_letter(last_year_col)}6*'Co. Desc'!{get_column_letter(last_year_col)}16"
        data_cell = ws.cell(row=40, column=4, value=formula)
    else:
        # Fallback to original code if no years found
        val = anip.get("net_income")
        if val is not None:
            val = val / 1_000_000
        data_cell = ws.cell(row=40, column=4, value=val)
    
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font

    ws.cell(row=41, column=3, value="Addback:").font = label_font
    val = anip.get("addback")
    if val is not None:
        val = val / 1_000_000
    data_cell = ws.cell(row=41, column=4, value=val)
    data_cell.number_format = '#,##0'
    data_cell.font = data_arial_font
    ws.cell(row=41, column=5, value="Merger charges, writedowns above the line, dep. Amort below the line less capex").font = data_arial_font

    # Years Payback - now using formula =D39/(D40+D41) instead of static value
    ws.cell(row=42, column=3, value="Years Payback:").font = label_font
    formula = "=D39/(D40+D41)"
    data_cell = ws.cell(row=42, column=4, value=formula)
    data_cell.number_format = '#,##0.0'
    data_cell.font = data_arial_font
    
    # Apply data fills to specified ranges
    def data_fill_range(ws, top_row, left_col, bottom_row, right_col):
        for rr in range(top_row, bottom_row + 1):
            for cc in range(left_col, right_col + 1):
                cell = ws.cell(rr, cc)
                if cell.fill.patternType is None:
                    cell.fill = data_fill

    data_fill_range(ws, 3, 2, 5, 12)   # B3:L5
    data_fill_range(ws, 7, 2, 9, 12)   # B7:L9
    data_fill_range(ws, 15, 2, 17, 12) # B15:L17
    data_fill_range(ws, 23, 2, 25, 12) # B23:L25
    data_fill_range(ws, 35, 2, 37, 12) # B35:L37

def to_float(val):
    if val is None or val == "":
        return None
    val_str = str(val).replace("%", "").replace(",", "").strip()
    if val_str == "":
        return None
    try:
        return float(val_str)
    except ValueError:
        return None
    
def write_qualities_sheet(writer, final_output):
    """
    Create or update a sheet called 'Qualities' that displays the text from
    final_output['qualities'] with:
    - Numbered items (1-10)
    - Bold headers inline with numbers
    - Wrapped descriptive text starting on the next line
    Each quality is separated by a blank line for readability.
    """
    if not final_output.get("qualities"):
        return
    
    wb = writer.book

    # Create sheet if it doesn't exist
    if "Qualities" not in wb.sheetnames:
        wb.create_sheet("Qualities")
    ws = wb["Qualities"]

    # Get the qualities text
    text = final_output.get("qualities", "No summary available.")

    # Set title
    ws["A1"] = "Core Analysis"
    ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
    ws["A1"].fill = label_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = thin_border

    # Split into individual quality entries (split on numbered items)
    qualities = re.split(r'\n\n(?=\d+\.)', text.strip())
    
    current_row = 3
    col = 1

    for quality in qualities:
        if not quality.strip():
            continue
            
        # Extract the number, header, and description
        # Pattern matches: number, bold header, and description
        match = re.match(r'(\d+)\.\s*\*\*(.*?)\*\*:(.+)', quality.strip(), re.DOTALL)
        
        if match:
            number, header, description = match.groups()
            
            # Write the numbered header line
            header_cell = ws.cell(row=current_row, column=col)
            header_cell.value = f"{number}. {header}:"
            header_cell.font = Font(name="Arial", size=10, bold=True)
            current_row += 1
            
            # Write the wrapped description on subsequent lines
            description = description.strip()
            wrapped_lines = textwrap.wrap(description, width=100)
            
            for line in wrapped_lines:
                desc_cell = ws.cell(row=current_row, column=col)
                desc_cell.value = line
                desc_cell.font = Font(name="Arial", size=10)
                current_row += 1
            
            # Add a blank line between qualities
            current_row += 1

    # Set column width
    ws.column_dimensions[get_column_letter(col)].width = 110

    # Remove gridlines
    ws.sheet_view.showGridLines = False

    # Optionally set column width so text fits nicely
    ws.column_dimensions[get_column_letter(col)].width = 110

    # Remove gridlines on this sheet if you prefer
    ws.sheet_view.showGridLines = False

def write_industry_sheet(writer, final_output):
    """
    Write the Industry sheet with operating and market statistics.
    For the first company's Operating Margin, references the most recent year's value in the Analyses sheet.
    
    Parameters:
    writer: ExcelWriter object
    final_output: Dictionary containing the full output data including industry statistics
    """
    # Early return if qualities is None/null
    if not final_output.get("industry"):
        return
    
    wb = writer.book
    
    # If the sheet doesn't exist yet, create it
    if "Industry" not in wb.sheetnames:
        wb.create_sheet("Industry")
    ws = wb["Industry"]

    industry_data = final_output["industry"]
    industry_name = final_output["summary"]["industry"]
    # Write and format the "Industry Overview" title
    title_cell = ws.cell(row=1, column=6, value=f"Industry Comparison: {industry_name}")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    apply_table_border(ws, 1, 4, 8)
    title_fill_range(ws, 1, 4, 8)

    # Write Operating Statistics section
    op_stats_cell = ws.cell(row=3, column=2, value="Operating Statistics:")
    op_stats_cell.fill = label_fill
    op_stats_cell.font = label_font
    op_stats_cell.border = thin_border
    
    # Get companies (tickers)
    companies = list(industry_data["operatingStatistics"].keys())
    
    # Define the operating statistics columns and their formats
    op_stats_columns = {
        "Company": (2, None),  # Column B, no special format
        "Debt(yrs.)": (4, '#,##0.0'),  # Column D
        "Sales": (6, '#,##0'),  # Column F
        "ROC": (8, '0.0%'),  # Column H
        "Operating Margin": (10, '0.0%')  # Column J
    }

    # Write operating statistics headers and data
    row = 5  # Start at row 5
    
    # Write headers
    for label, (col, _) in op_stats_columns.items():
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = label_fill
        cell.font = label_font
        cell.border = thin_border
        cell.alignment = center_alignment

    # Get the most recent year from the Analyses sheet for operating margin reference
    most_recent_year_col = None
    if "Analyses" in wb.sheetnames:
        analyses_ws = wb["Analyses"]
        
        # Find the last year column in the Analyses sheet
        # Years are in row 9, starting from column B (2)
        col = 2
        while True:
            year_cell = analyses_ws.cell(row=9, column=col)
            if year_cell.value is None or col > analyses_ws.max_column:
                break
            most_recent_year_col = col
            col += 1
    
    # Write company data
    for idx, company in enumerate(companies):
        row += 1
        # Write company name
        company_cell = ws.cell(row=row, column=2, value=company)
        company_cell.font = data_arial_font
        company_cell.alignment = center_alignment
        company_data = industry_data["operatingStatistics"][company]
        
        # Write operating statistics
        col_mappings = {
            "Debt(yrs.)": (4, "Debt(yrs.)"),
            "Sales": (6, "Sales"),
            "ROC": (8, "ROC"),
            "Operating Margin": (10, "Operating Margin")
        }

        for label, (col, key) in col_mappings.items():
            # Special handling for first company's debt years and operating margin
            if idx == 0:
                if label == "Debt(yrs.)":
                    cell = ws.cell(row=row, column=col, value="='Studies'!D42")
                elif label == "Operating Margin" and most_recent_year_col is not None:
                    # Reference the most recent year's operating margin in Analyses sheet
                    cell = ws.cell(row=row, column=col, value=f"='Analyses'!{get_column_letter(most_recent_year_col-2)}20")
                else:
                    value = company_data[key]
                    cell = ws.cell(row=row, column=col, value=value)
                    
                    # Convert Sales to millions
                    if key == "Sales":
                        cell.value = value / 1_000_000
            else:
                value = company_data[key]
                cell = ws.cell(row=row, column=col, value=value)
                
                # Convert Sales to millions
                if key == "Sales":
                    cell.value = value / 1_000_000
            
            cell.font = data_arial_font
            cell.alignment = center_alignment
            cell.number_format = op_stats_columns[label][1]

    last_op_stats_row = row

    # Market Statistics section
    market_stats_start_row = last_op_stats_row + 3
    market_stats_cell = ws.cell(row=market_stats_start_row, column=2, value="Market Statistics:")
    market_stats_cell.fill = label_fill
    market_stats_cell.font = label_font
    market_stats_cell.border = thin_border

    # Define the market statistics columns and their formats
    market_stats_columns = {
        "Company": (2, None),  # Column B, no special format
        "P/B": (4, '#,##0.00'),  # Column D
        "P/E": (6, '#,##0.0'),  # Column F
        "Div. Yld.": (8, '0.00%'),  # Column H
        "EV/Sales": (10, '#,##0.00')  # Column J
    }

    # Write market statistics headers
    row = market_stats_start_row + 2
    for label, (col, _) in market_stats_columns.items():
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = label_fill
        cell.font = label_font
        cell.border = thin_border
        cell.alignment = center_alignment

    # Write market statistics data
    for company in companies:
        row += 1
        # Write company name
        company_cell = ws.cell(row=row, column=2, value=company)
        company_cell.font = data_arial_font
        company_cell.alignment = center_alignment
        company_data = industry_data["marketStatistics"][company]
        
        # Write market statistics
        col_mappings = {
            "P/B": (4, "P/B"),
            "P/E": (6, "P/E"),
            "Div. Yld.": (8, "Div. Yld."),
            "EV/Sales": (10, "EV/Sales")
        }

        for label, (col, key) in col_mappings.items():
            value = company_data[key]
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_arial_font
            cell.alignment = center_alignment
            cell.number_format = market_stats_columns[label][1]

    # Adjust column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 15

def write_hist_pricing_sheet(writer, final_output):
    """
    Write the Historical Pricing sheet with average ratios and price implications in a 2x2 grid layout
    Using formulas to calculate average low and high ratios based on historical data
    
    Parameters:
    writer: ExcelWriter object
    final_output: Dictionary containing the full output data including historical pricing and current metrics
    """
    wb = writer.book
    
    if "Hist. Pricing" not in wb.sheetnames:
        wb.create_sheet("Hist. Pricing")
    ws = wb["Hist. Pricing"]

    hist_pricing = final_output.get("historical_pricing", {})
    reported_currency = final_output["summary"]["reported_currency"]

    # Write and format the title
    title_cell = ws.cell(row=1, column=4, value=f"Historical Pricing Analysis ({reported_currency})")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    apply_table_border(ws, 1, 3, 6)
    title_fill_range(ws, 1, 3, 6)

    # Get all years from Co. Desc sheet for formulas
    sorted_years = sorted(final_output["company_description"]["data"].keys(), key=lambda x: int(x))
    if sorted_years:
        max_year = max(int(year) for year in sorted_years)
        new_year = str(max_year + 1)
        # In Co. Desc sheet, years start at column B (2)
        first_new_year_col = get_column_letter(2 + len(sorted_years))
    else:
        first_new_year_col = "B"  # fallback
    
    # Get the range of historical years for our average calculations
    first_hist_year_col = get_column_letter(2)  # Column B in Co. Desc
    last_hist_year_col = get_column_letter(2 + len(sorted_years) - 1)  # Last historical year column

    # Define the grid positions for each metric
    metrics = {
        # Top Left - P/E Ratio
        "P/E Ratio": {
            "low_formula": f"=AVERAGE(('Co. Desc'!{first_hist_year_col}9:{last_hist_year_col}9)/('Co. Desc'!{first_hist_year_col}6:{last_hist_year_col}6))",
            "high_formula": f"=AVERAGE(('Co. Desc'!{first_hist_year_col}10:{last_hist_year_col}10)/('Co. Desc'!{first_hist_year_col}6:{last_hist_year_col}6))",
            "current_formula": f"='Co. Desc'!{first_new_year_col}5",  # References diluted EPS
            "start_row": 3,
            "start_col": 2,
            "format": '#,##0.0',
            "value_type": "earnings"
        },
        # Top Right - P/S Ratio
        "P/S Ratio": {
            "low_formula": f"=AVERAGE(('Co. Desc'!{first_hist_year_col}9:{last_hist_year_col}9)/('Analyses'!{first_hist_year_col}11:{last_hist_year_col}11))",
            "high_formula": f"=AVERAGE(('Co. Desc'!{first_hist_year_col}10:{last_hist_year_col}10)/('Analyses'!{first_hist_year_col}11:{last_hist_year_col}11))",
            "current_formula": f"='Analyses'!{first_new_year_col}11",  # References Sales/Share
            "start_row": 3,
            "start_col": 8,
            "format": '#,##0.00',
            "value_type": "sales"
        },
        # Bottom Left - P/B Ratio
        "P/B Ratio": {
            "low_formula": f"=AVERAGE(('Co. Desc'!{first_hist_year_col}9:{last_hist_year_col}9)/('Co. Desc'!{first_hist_year_col}20:{last_hist_year_col}20))",
            "high_formula": f"=AVERAGE(('Co. Desc'!{first_hist_year_col}10:{last_hist_year_col}10)/('Co. Desc'!{first_hist_year_col}20:{last_hist_year_col}20))",
            "current_formula": f"='Co. Desc'!{first_new_year_col}20",  # References Book Value/Share
            "start_row": 10,
            "start_col": 2,
            "format": '#,##0.00',
            "value_type": "book_value"
        },
        # Bottom Right - P/CF Ratio
        "P/CF Ratio": {
            "low_formula": f"=AVERAGE(('Co. Desc'!{first_hist_year_col}9:{last_hist_year_col}9)/(('Co. Desc'!{first_hist_year_col}4:{last_hist_year_col}4+'Analyses'!{first_hist_year_col}22:{last_hist_year_col}22)/('Co. Desc'!{first_hist_year_col}16:{last_hist_year_col}16)))",
            "high_formula": f"=AVERAGE(('Co. Desc'!{first_hist_year_col}10:{last_hist_year_col}10)/(('Co. Desc'!{first_hist_year_col}4:{last_hist_year_col}4+'Analyses'!{first_hist_year_col}22:{last_hist_year_col}22)/('Co. Desc'!{first_hist_year_col}16:{last_hist_year_col}16)))",
            "current_formula": f"=('Co. Desc'!{first_new_year_col}4+'Analyses'!{first_new_year_col}22)/'Co. Desc'!{first_new_year_col}16",  # (Net Profit + Depreciation) / Shares Outstanding
            "start_row": 10,
            "start_col": 8,
            "format": '#,##0.00',
            "value_type": "cash_flow"
        }
    }

    # Write each metric box
    for metric, props in metrics.items():
        start_row = props["start_row"]
        start_col = props["start_col"]
        
        # Write box title
        title_cell = ws.cell(row=start_row, column=start_col, value=metric)
        title_cell.fill = label_fill
        title_cell.font = label_font
        title_cell.alignment = center_alignment
        ws.merge_cells(start_row=start_row, start_column=start_col, 
                      end_row=start_row, end_column=start_col + 1)
        
        # Write metric rows vertically
        metrics_data = [
            ("Used", props["current_formula"]),
            ("Avg Low", props["low_formula"]),
            ("Avg High", props["high_formula"])
        ]
        
        for idx, (label, formula) in enumerate(metrics_data):
            # Write label
            label_cell = ws.cell(row=start_row + 1 + idx, column=start_col, value=label)
            label_cell.fill = label_fill
            label_cell.font = label_font
            label_cell.border = thin_border
            label_cell.alignment = center_alignment
            
            # Write formula
            value_cell = ws.cell(row=start_row + 1 + idx, column=start_col + 1, value=formula)
            value_cell.fill = data_fill
            value_cell.font = data_arial_font
            value_cell.border = thin_border
            value_cell.number_format = props["format"]
            value_cell.alignment = right_alignment
        
        # Write Buy and Sell rows with formulas
        used_cell = f"{get_column_letter(start_col + 1)}{start_row + 1}"
        avg_low_cell = f"{get_column_letter(start_col + 1)}{start_row + 2}"
        avg_high_cell = f"{get_column_letter(start_col + 1)}{start_row + 3}"
        
        # Write Buy row (Used * Avg Low)
        buy_label = ws.cell(row=start_row + 4, column=start_col, value="Buy")
        buy_label.fill = label_fill
        buy_label.font = label_font
        buy_label.border = thin_border
        buy_label.alignment = center_alignment
        
        buy_cell = ws.cell(row=start_row + 4, column=start_col + 1, 
                          value=f"={used_cell}*{avg_low_cell}")
        buy_cell.fill = data_fill
        buy_cell.font = data_arial_bold_font
        buy_cell.border = thin_border
        buy_cell.number_format = '#,##0.00'
        buy_cell.alignment = right_alignment
        
        # Write Sell row (Used * Avg High)
        sell_label = ws.cell(row=start_row + 5, column=start_col, value="Sell")
        sell_label.fill = label_fill
        sell_label.font = label_font
        sell_label.border = thin_border
        sell_label.alignment = center_alignment
        
        sell_cell = ws.cell(row=start_row + 5, column=start_col + 1, 
                           value=f"={used_cell}*{avg_high_cell}")
        sell_cell.fill = data_fill
        sell_cell.font = data_arial_bold_font
        sell_cell.border = thin_border
        sell_cell.number_format = '#,##0.00'
        sell_cell.alignment = right_alignment
        
        # Add border around the entire box
        for row in range(start_row, start_row + 6):
            for col in range(start_col, start_col + 2):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

    # Adjust column widths
    for base_col in [2, 8]:  # Starting columns for left and right sections
        ws.column_dimensions[get_column_letter(base_col)].width = 12      # Labels
        ws.column_dimensions[get_column_letter(base_col + 1)].width = 20  # Values

def write_valuation_sheet(writer, final_output, ticker):
    """Write the valuation analysis sheet with 2x2 grid layout"""
    reported_currency = final_output["summary"]["reported_currency"]
    wb = writer.book
    
    if "Valuation" not in wb.sheetnames:
        wb.create_sheet("Valuation")
    ws = wb["Valuation"]

    # Title
    title_cell = ws.cell(row=1, column=4, value="Valuation (USD)")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_fill_range(ws, 1, 3, 5)
    apply_table_border(ws, 1, 3, 5)

    current_price = get_current_quote_yahoo(ticker)

    # Get forecast year column reference
    sorted_years = sorted(
        final_output["company_description"]["data"].keys(),
        key=lambda x: int(x)
    )
    first_forecast_col = (
        get_column_letter(2 + len(sorted_years)) if sorted_years else "B"
    )

    # =========================================================================
    # REARRANGED SETTINGS
    # =========================================================================
    #   B3 => ADR Multiple
    #   B4 => Currency Ratio
    #   D3 => EPS Growth
    #   D4 => Dividend Growth
    #   F3 => Purchase Discount
    #   F4 => Sell Discount
    #   F5 => PE Multiple
    # =========================================================================
    settings = {
        (3, 2): ("ORD:ADR:", 1),                           # B3
        (4, 2): (f"USD:{reported_currency} rate:", 1),          # B4
        (3, 4): ("EPS growth rate:", 0.10),                     # D3
        (4, 4): ("Dividend growth rate:", 0.10),                # D4
        (3, 6): ("Purchase Discount:", 0.14),                   # F3
        (4, 6): ("Sell Discount:", 0.05),                       # F4
        (5, 6): ("PE Multiple:", 25),                           # F5
    }

    for (row, col), (label, value) in settings.items():
        label_cell = ws.cell(row=row, column=col - 1, value=label)
        label_cell.fill = label_fill
        label_cell.font = label_font
        label_cell.border = thin_border
        
        value_cell = ws.cell(row=row, column=col, value=value)
        value_cell.fill = data_fill
        value_cell.font = data_arial_font
        value_cell.border = thin_border
        value_cell.alignment = center_alignment

        # Format anything < 1 as percentage
        if isinstance(value, float) and value < 1:
            value_cell.number_format = '0.00%'
    
    new_settings = {
        (3, 7): ("Buy %:", 0.60),    # G3 (label), H3 (value = 60%)
        (4, 7): ("Sell %:", 1.20),   # G4 (label), H4 (value = 120%)
    }

    for (row, col), (label, value) in new_settings.items():
        label_cell = ws.cell(row=row, column=col, value=label)
        label_cell.fill = label_fill
        label_cell.font = label_font
        label_cell.border = thin_border
        
        value_cell = ws.cell(row=row, column=col + 1, value=value)
        value_cell.fill = data_fill
        value_cell.font = data_arial_font
        value_cell.border = thin_border
        value_cell.alignment = center_alignment

        # Format them as percentages
        value_cell.number_format = '0.00%'

    # If you need the numeric value of PE multiple in code:
    # pe_multiple_val = ws.cell(row=5, column=6).value  # F5

    # =========================================================================
    # 2x2 GRID LAYOUT
    # =========================================================================
    # We keep top-left and bottom-left in columns B/C.
    # We move top-right and bottom-right to columns E/F.
    # Then fix all formula references accordingly.
    # =========================================================================
    grid_segments = {
        # ---------------------------------------------------------------------
        # TOP LEFT (columns B/C)
        # ---------------------------------------------------------------------
        "Initial Rate of Investment:": {
            "start_row": 8,
            "start_col": 2,  # B
            "metrics": {
                "Current Price:": f"={current_price}",
                # Currency ratio is B4, ADR multiple is B3:
                "Current EPS:": f"='Co. Desc'!{first_forecast_col}5 * B4 * B3",
                # = B10 / B9 once written to the sheet
                "Initial ROI:": "=B10/B9",
            },
        },
        # ---------------------------------------------------------------------
        # TOP RIGHT (columns E/F)
        # ---------------------------------------------------------------------
        "Relative Value to Investment In T-Bonds:": {
            "start_row": 8,
            "start_col": 5,  # E
            "metrics": {
                # Same logic as top-left for 'Current EPS'
                "Current EPS:": f"='Co. Desc'!{first_forecast_col}5 * B4 * B3",
                "T-Bond Rate:": 0.04,
                # Was =H9/H10 in original; now =E9/E10
                "Relative Value:": "=E9/E10",
            },
        },
        # ---------------------------------------------------------------------
        # BOTTOM LEFT (columns B/C)
        # ---------------------------------------------------------------------
        "Valuation as an Equity Bond:": {
            "start_row": 15,
            "start_col": 2,  # B
            "metrics": {
                # Use B3 (ADR multiple) and B4 (currency ratio)
                "Current BV:": f"='Co. Desc'!{first_forecast_col}20 * B3 * B4",
                "Current ROE:": f"='Co. Desc'!{first_forecast_col}24",
                "Retained % adjustment:": 0.10,
                # unchanged, presumably references other sheet cells
                "Retained %:": "=1 - 'Analyses'!K5 - 'Analyses'!K7 - B18",
                "Net BV growth:": "=B17*B19",
                "BV in year 10:": "=FV(B20, 10, , -B16)",
                "EPS Adjustment Factor:": 1.5,
                "EPS in Year 10:": "=B17 * B21 * B22",
                # PE multiple is at F5
                "Value at PE Multiple:": "=F5 * B23",
                # Dividend growth is at D4 ⇒ use FV(D4,10,…)
                "Total Dividends:": (
                    f"=(('Co. Desc'!{first_forecast_col}13 + "
                    f"FV(D4, 10, , -'Co. Desc'!{first_forecast_col}13))/2)*10*B3*B4"
                ),
                "Total Future Value:": "=B24+B25",
                # Purchase discount is at F3
                "Purchase at Discount:": "=PV(F3, 10, , B26)*-1",
            },
        },
        # ---------------------------------------------------------------------
        # BOTTOM RIGHT (columns E/F)
        # ---------------------------------------------------------------------
        "Valuation on Earnings Growth:": {
            "start_row": 15,
            "start_col": 5,  # E
            "metrics": {
                # Same B3/B4 for ADR/currency
                "Current EPS:": f"='Co. Desc'!{first_forecast_col}5 * B4 * B3",
                # EPS Growth is at D3 => FV(D3,10,…)
                "EPS in year 10:": "=FV(D3, 10, , -E16)",
                "Avg PE Ratio:": f"=AVERAGE('Co. Desc'!B8:{first_forecast_col}8)",
                # Was =B5*H17 + B25; now =F5*E17 + B25
                "Value at PE Multiple:": "=F5*E17 + B25",
                # Was =RATE(10, , B9, -H19 + B25); now =RATE(10, , B9, -E19 + B25)
                "Price Return:": "=RATE(10, , B9, -E19 + B25)",
                "Dividend Return:": f"='Co. Desc'!{first_forecast_col}14",
                # Was =H20+H21; now =E20+E21
                "Total Return:": "=E20 + E21",
                # Purchase discount is F3
                "Purchase at Discount:": "=PV(F3, 10, , -E19)",
                # Sell discount is F4
                "Sell at Discount:": "=PV(F4, 10, , -E19)",
            },
        },
        "Capital Charge Approach:": {
            "start_row": 8,
            "start_col": 8,  # G
            "metrics": {
                # EBIT = operating margin * sales (most recent year)
                #   If your Co. Desc sheet has operating margin in row 17 and
                #   sales in row 12 for the "most recent year," adapt as needed.
                "EBIT:": f"='Analyses'!{first_forecast_col}20 * 'Analyses'!{first_forecast_col}10 * B4",
                
                # Required return (cost of capital) => default 12%
                "Required Return:": 0.12,
                
                # Growth rate => reference the EPS growth rate in D3 if you like
                "Growth Rate:": "=D3",
                
                # % not required => default 80%
                "% Not Required:": 0.80,
                
                # Denominator = required return - (% not required * growth rate)
                #   Will reference the corresponding cells once placed:
                #   If "Required Return:" goes in H10,
                #   and "Growth Rate:" in H11,
                #   and "% Not Required:" in H12,
                #   then Denominator = H10 - (H11 * H12).
                "Denominator:": "=H10 - (H11 * H12)",
                
                # EV = EBIT / Denominator  (EBIT in H8, Denominator in H12)
                #   NOTE: watch the row offsets as you fill out the items below!
                "EV:": "=H9 / H13",
                
                # Debt => from 'Studies'!D39 (Long-term debt)
                "Debt:": "='Studies'!D39 * B4",
                
                # Equity Value = EV - Debt
                "Equity Value:": "=(H14 - H15)",
                
                # Shares Outstanding => from co desc, row ? 
                #   If your co desc has shares in row 2 for that "year," adapt as needed.
                "Shares Outstanding:": f"='Co. Desc'!{first_forecast_col}16 * (1/B3)",
                
                # Share Value = Equity Value / Shares Outstanding
                "Share Value:": "=H16 / H17",
                
                # Buy at => Share Value * Buy % (which is in H1)
                "Buy At:": "=H18 * $H$3",
                
                # Sell at => Share Value * Sell % (which is in H2)
                "Sell At:": "=H18 * $H$4",
            },
        }
    }

    #
    # Write each 2×2 grid segment
    #
    for title, config in grid_segments.items():
        start_row = config["start_row"]
        start_col = config["start_col"]
        
        # Segment title
        title_cell = ws.cell(row=start_row, column=start_col - 1, value=title)
        title_cell.fill = label_fill
        title_cell.font = Font(name="Times New Roman", size=12, bold=True, italic=True)
        title_cell.border = thin_border
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Merge the two columns for the label
        ws.merge_cells(
            start_row=start_row, start_column=start_col - 1, 
            end_row=start_row, end_column=start_col
        )
        
        # Write metrics
        current_row = start_row + 1
        for label, formula in config["metrics"].items():
            # Label
            label_cell = ws.cell(row=current_row, column=start_col - 1, value=label)
            label_cell.fill = label_fill
            label_cell.font = label_font
            label_cell.border = thin_border
            
            # Value/Formula
            value_cell = ws.cell(row=current_row, column=start_col, value=formula)
            value_cell.fill = data_fill
            value_cell.font = data_arial_font
            value_cell.border = thin_border
            value_cell.alignment = center_alignment

            # Make certain items bold
            if label.lower() in ["relative value:", "purchase at discount:", "sell at discount:", "buy at:", "sell at:"]:
                value_cell.font = data_arial_bold_font
            
            # Format numeric cells
            label_lower = label.lower()
            if label_lower == "t-bond rate:":
                value_cell.number_format = '0.00%'
            elif label_lower == "eps adjustment factor:":
                value_cell.number_format = '0.00'
            elif label_lower == "avg pe ratio:":
                value_cell.number_format = '0.00'
            elif "net bv growth" in label_lower:
                value_cell.number_format = '0.00%'
            elif "shares outstanding" in label_lower:
                value_cell.number_format = '#,##0'
            elif any(x in label_lower for x in ["rate", "roi", "return", "roe", "%"]):
                value_cell.number_format = '0.00%'
            elif any(x in label_lower for x in ["price", "value", "eps", "bv", "dividends", "purchase", "sell",
                                                "ebit", "value", "debt", "share", "buy at", "sell at"]):
                value_cell.number_format = '"$"#,##0.00'
            
            current_row += 1

    # =========================================================================
    # Column Widths
    # =========================================================================
    # Example adjustments for label columns vs. value columns
    ws.column_dimensions[get_column_letter(1)].width = 25  # A
    ws.column_dimensions[get_column_letter(2)].width = 15  # B
    ws.column_dimensions[get_column_letter(3)].width = 25  # C
    ws.column_dimensions[get_column_letter(4)].width = 20  # D
    ws.column_dimensions[get_column_letter(5)].width = 25  # E
    ws.column_dimensions[get_column_letter(6)].width = 15  # F
    ws.column_dimensions[get_column_letter(7)].width = 25  # G
    ws.column_dimensions[get_column_letter(8)].width = 15  # H

def write_segmentation_sheet(writer, final_output):
    """
    Write the Segmentation sheet showing revenue breakdown by business segment over time
    
    Parameters:
    writer: ExcelWriter object
    final_output: Dictionary containing the full output data including segmentation data
    """
    # Early return if segmentation data is not present
    if not final_output.get("segmentation"):
        return
    
    reported_currency = final_output["summary"]["reported_currency"]
    segmentation_data = final_output["segmentation"]
    wb = writer.book
    
    # Create the sheet if it doesn't exist
    if "Segmentation" not in wb.sheetnames:
        wb.create_sheet("Segmentation")
    ws = wb["Segmentation"]
    ws.freeze_panes = "B1"

    # Write and format the title
    title_cell = ws.cell(row=1, column=5, value=f"Revenue Segmentation (in mlns {reported_currency})")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    apply_table_border(ws, 1, 3, 7)
    title_fill_range(ws, 1, 3, 7)

    # Get sorted years and all unique segments
    sorted_years = sorted(segmentation_data.keys(), key=lambda x: int(x))
    all_segments = set()
    for year_data in segmentation_data.values():
        all_segments.update(year_data.keys())
    sorted_segments = sorted(all_segments)

    # Write year headers starting at row 3
    for i, year in enumerate(sorted_years):
        year_col = i + 2  # Start at column B
        y_cell = ws.cell(row=3, column=year_col, value=year)
        y_cell.fill = label_fill
        y_cell.font = label_font
        y_cell.border = thin_border
        y_cell.alignment = center_alignment

    # Write segment data
    current_row = 4
    for segment in sorted_segments:
        # Write segment name in column A
        segment_cell = ws.cell(row=current_row, column=1, value=segment)
        segment_cell.font = label_font
        segment_cell.fill = label_fill
        segment_cell.border = thin_border

        # Write values for each year
        for i, year in enumerate(sorted_years):
            year_col = i + 2  # Start at column B
            value = segmentation_data[year].get(segment)
            
            # Create cell and apply styling regardless of value
            value_cell = ws.cell(row=current_row, column=year_col)
            value_cell.fill = data_fill
            value_cell.font = data_arial_font
            value_cell.border = thin_border
            value_cell.alignment = right_alignment
            
            if value is not None:
                # Convert to millions and set value
                value = value / 1_000_000
                value_cell.value = value
                value_cell.number_format = '#,##0'

        current_row += 1

    # Calculate and write growth rates in the rightmost column
    growth_col = len(sorted_years) + 3  # One column after percentages
    growth_header = ws.cell(row=3, column=growth_col, value="CAGR")
    growth_header.fill = label_fill
    growth_header.font = label_font
    growth_header.border = thin_border
    growth_header.alignment = center_alignment

    for segment_idx, segment in enumerate(sorted_segments):
        row = segment_idx + 4
        
        # Get first and last valid values
        first_val = next((segmentation_data[year].get(segment) for year in sorted_years if segmentation_data[year].get(segment) is not None), None)
        last_val = next((segmentation_data[year].get(segment) for year in reversed(sorted_years) if segmentation_data[year].get(segment) is not None), None)
        
        if first_val and last_val and first_val != 0:
            years_between = len(sorted_years) - 1
            if years_between > 0:
                cagr = (last_val / first_val) ** (1/years_between) - 1
                growth_cell = ws.cell(row=row, column=growth_col, value=cagr)
                growth_cell.number_format = '0.0%'
                growth_cell.font = Font(name="Arial", size=8, italic=True)
                growth_cell.fill = data_fill  # Added fill for consistency
                growth_cell.border = thin_border  # Added border for consistency

        else:
            # Create empty growth cell with consistent styling
            growth_cell = ws.cell(row=row, column=growth_col)
            growth_cell.fill = data_fill
            growth_cell.border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 30  # Segment names
    for col in range(2, growth_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

def sync_operating_margin_from_profit_desc(writer):
    """
    Updates the operating margin row in the Analyses sheet with formula references 
    to the corresponding percentage cells in the Profit.Desc. sheet.
    
    This function replaces static values or existing formulas with direct references
    to ensure both sheets always show the same operating margin percentages.
    
    Parameters:
    writer: ExcelWriter object with access to the workbook
    """
    wb = writer.book
    
    # Ensure both required sheets exist
    if "Analyses" not in wb.sheetnames or "Profit.Desc." not in wb.sheetnames:
        print("Warning: Cannot sync operating margin - required sheets not found")
        return
    
    analyses_ws = wb["Analyses"]
    profit_desc_ws = wb["Profit.Desc."]
    
    # Find the operating margin row in analyses sheet (should be row 20)
    operating_margin_row = 20  # Default from the original code
    
    # Find the operating margin row and year columns in profit_desc sheet
    operating_margin_row_pd = None
    op_margin_label = "Operating Margin:"
    
    # Find the row with "Operating Margin:" label in column A
    for row in range(1, profit_desc_ws.max_row + 1):
        cell_value = profit_desc_ws.cell(row=row, column=1).value
        if cell_value == op_margin_label:
            operating_margin_row_pd = row
            break
    
    if operating_margin_row_pd is None:
        print("Warning: 'Operating Margin:' row not found in Profit.Desc. sheet")
        return
    
    # Find the year columns in both sheets
    # In Analyses, years start at column B (index 2)
    # In Profit.Desc, years start at column D (index 4) and appear every 2 columns
    
    # Get years from Analyses sheet
    analyses_years = []
    col = 2  # Starting at column B
    while True:
        year_cell = analyses_ws.cell(row=9, column=col)  # Where years are in Analyses
        if year_cell.value is None:
            break
        analyses_years.append((col, str(year_cell.value)))
        col += 1
    
    # Get years from Profit.Desc sheet
    profit_desc_years = []
    col = 4  # Starting at column D
    while True:
        year_cell = profit_desc_ws.cell(row=3, column=col)  # Where years are in Profit.Desc
        if year_cell.value is None or col > profit_desc_ws.max_column:
            break
        profit_desc_years.append((col, str(year_cell.value)))
        col += 2  # Skip percentage column
    
    # Match years and create formula references
    for a_col, a_year in analyses_years:
        for pd_col, pd_year in profit_desc_years:
            if a_year == pd_year:
                # Found matching year
                # Create a reference to the percentage cell (column to the right of value in Profit.Desc)
                target_col = pd_col + 1  # Percentage column is one to the right of the year column
                
                # Create formula referencing the percentage cell in Profit.Desc
                formula = f"='Profit.Desc.'!{get_column_letter(target_col)}{operating_margin_row_pd}"
                
                # Update the cell in Analyses with the formula
                cell = analyses_ws.cell(row=operating_margin_row, column=a_col, value=formula)
                
                # Format as percentage
                cell.number_format = '0.0%'
                break
    
    print("Successfully added operating margin formula references from Profit.Desc. to Analyses sheet")

def generate_config_note(ticker, wb):
    """
    Add a note in cell B1 of the profit_desc sheet if there are any
    configuration overrides for the company from financial_data_config.json.
    
    Args:
        ticker (str): Company ticker symbol
        wb (openpyxl.Workbook): Excel workbook object
    """
    import json
    import logging
    from pathlib import Path

    logger = logging.getLogger(__name__)
    
    # Get the profit_desc sheet
    try:
        sheet = wb["Profit.Desc."]
    except KeyError:
        logger.warning("profit_desc sheet not found in workbook")
        return
    
    # Load the configuration file
    try:
        with open('financial_data_config.json', 'r') as f:
            config = json.load(f)
    except FileNotFoundError:
        logger.warning("financial_data_config.json not found")
        return
    except json.JSONDecodeError:
        logger.warning("Error parsing financial_data_config.json")
        return
    
    # Check if we have any configurations for this ticker
    ticker_config = config.get(ticker.upper(), {})
    if not ticker_config:
        return
    
    # Build the configuration note
    notes = []
    for statement, fields in ticker_config.items():
        field_notes = []
        for field, value in fields.items():
            field_notes.append(f"set {field} to {value}")
        if field_notes:
            notes.append(f"In {statement}, {', '.join(field_notes)}")
    
    if notes:
        note = '. '.join(notes)
        # Add note to cell B1
        sheet['A2'] = f"Configuration overrides: {note}."
        
        # Style the cell
        from openpyxl.styles import Font, PatternFill
        cell = sheet['B1']
        cell.font = Font(italic=True, size=9, color="666666")
        cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

def generate_excel_for_ticker_year(ticker: str, year: int, no_add_da: bool = False):
    """
    Generate the Excel file for the given ticker and year, writing to:
       ./output/{ticker}.{last 2 digits of year}.2.xlsx
    
    :param ticker: The company symbol/ticker.
    :param year:   The 4-digit year (e.g., 2024). We'll use only the last two digits in the filename.
    """
    # Convert to uppercase and build the filename, e.g.: ./output/ABC.24.2.xlsx
    ticker = ticker.upper()
    year_2_digits = str(year)[-2:]  # last two chars of the year
    xls_filename = os.path.join("output", f"{ticker}.{year_2_digits}.2.xlsx")

    # 1. Load the JSON data
    final_output = load_final_output(ticker)

    # 2. Create or append to the Excel file
    writer = create_xls(xls_filename)

    # 3. Write data to each sheet
    write_summary_sheet(writer, final_output)
    write_company_description(writer, final_output)
    write_analyses_sheet(writer, final_output)
    write_profit_desc_sheet(writer, final_output, no_add_da)
    
    write_balance_sheet_sheet(writer, final_output)
    write_studies_sheet(writer, final_output)
    write_qualities_sheet(writer, final_output)
    write_industry_sheet(writer, final_output)
    write_hist_pricing_sheet(writer, final_output)
    write_valuation_sheet(writer, final_output, ticker)
    write_segmentation_sheet(writer, final_output)
    generate_config_note(ticker, writer.book)

    sync_operating_margin_from_profit_desc(writer)

    # 4. Apply workbook formatting (remove gridlines, etc.)
    format_workbook(writer)

    # 5. Save
    writer.close()
    print(f"Data for {ticker} written to {xls_filename} successfully.")


if __name__ == "__main__":
    # Usage: python write_excel.py TICKER target_file.xls
    if len(sys.argv) < 3:
        print("Usage: python gen_excel.py TICKER XLS_FILENAME")
        sys.exit(1)

    ticker = sys.argv[1].upper()
    xls_filename = sys.argv[2]

    final_output = load_final_output(ticker)

    # Create or append to xls file
    writer = create_xls(xls_filename)

    # Now write data to each sheet
    write_summary_sheet(writer, final_output)
    write_company_description(writer, final_output)
    write_analyses_sheet(writer, final_output)
    write_profit_desc_sheet(writer, final_output)
    
    write_balance_sheet_sheet(writer, final_output)
    write_studies_sheet(writer, final_output)
    write_qualities_sheet(writer, final_output)
    write_industry_sheet(writer, final_output)
    write_hist_pricing_sheet(writer, final_output)
    write_valuation_sheet(writer, final_output, ticker)
    write_segmentation_sheet(writer, final_output)
    generate_config_note(ticker, writer.book)

    sync_operating_margin_from_profit_desc(writer)
    # Apply formatting: set font to Arial size 10 for non-formatted cells and remove gridlines
    format_workbook(writer)

    # Save changes
    writer.close()
    print(f"Data for {ticker} written to {xls_filename} successfully.")
