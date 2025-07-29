# gen_excel_bs.py

import os
import sys
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
import textwrap
import re

from utils import get_current_quote_yahoo, get_long_term_rate

# Define Custom Fills
label_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")  # Light blue
data_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")   # Cornsilk

# Define custom fonts
title_font = Font(name="Times New Roman", size=14, bold=True, italic=True)
label_font = Font(name="Times New Roman", size=10, bold=True, italic=True)
data_tnr_font = Font(name = "Times New Roman", size=10)
data_tnr_italic_font = Font(name='Times New Roman', size=10, italic=True)
data_tnr_bold_font = Font(name = "Times New Roman", bold=True)
data_arial_font = Font(name = "Arial", size=10)
data_arial_bold_font = Font(name ="Arial", size=10, bold=True)
data_arial_italic_font = Font(name = "Arial", size=10, italic=True)

center_alignment = Alignment(horizontal="center", vertical="center")
right_alignment = Alignment(horizontal="right", vertical="center")
# Define a thin black border
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Define a thick black border for the outer edges
thick_border = Border(
    left=Side(style='thick', color='000000'),
    right=Side(style='thick', color='000000'),
    top=Side(style='thick', color='000000'),
    bottom=Side(style='thick', color='000000')
)

def apply_table_border(ws, row, start_col, end_col):
    """
    Applies a thin border around a group of cells in a specified row from start_col to end_col.

    :param ws: The worksheet object.
    :param row: The row number where the group is located.
    :param start_col: The starting column number of the group.
    :param end_col: The ending column number of the group.
    """
    # Define a thin black border
    thin_side = Side(style='thin', color='000000')
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        existing_border = cell.border
        
        # Initialize new border components with existing borders
        new_border = Border(
            top=existing_border.top,
            bottom=existing_border.bottom,
            left=existing_border.left,
            right=existing_border.right,
            diagonal=existing_border.diagonal,
            diagonal_direction=existing_border.diagonal_direction,
            outline=existing_border.outline,
            vertical=existing_border.vertical,
            horizontal=existing_border.horizontal
        )
        
        # Apply top and bottom borders to all cells in the group
        new_border.top = thin_side
        new_border.bottom = thin_side
        
        # Apply left border to the first cell and right border to the last cell
        if col == start_col:
            new_border.left = thin_side
        if col == end_col:
            new_border.right = thin_side
        
        # Assign the updated border back to the cell
        cell.border = new_border

def title_fill_range(ws, row_number, left_col, right_col):
        for cc in range(left_col, right_col + 1):
            cell = ws.cell(row_number, cc)
            if cell.fill.patternType is None:
                cell.fill = label_fill

def format_workbook(writer):
    """
    removes gridlines from all worksheets.
    """
    for sheetname in writer.book.sheetnames:
        ws = writer.book[sheetname]
        # Remove gridlines
        ws.sheet_view.showGridLines = False
        
def load_final_output(ticker):
    file_path = os.path.join("output", f"{ticker}_yoy_consolidated_bs.json")
    if not os.path.exists(file_path):
        print(f"Error: {file_path} does not exist. Please generate {ticker}_yoy_consolidated_bs.json first.")
        sys.exit(1)
    with open(file_path, "r") as f:
        data = json.load(f)
    return data

def create_xls(xls_filename):
    if os.path.exists(xls_filename):
        print(f"Overwriting existing file: {xls_filename}")
    return pd.ExcelWriter(xls_filename, engine='openpyxl', mode='w')

def write_summary_sheet(writer, final_output):
    wb = writer.book

    if 'Summary' not in wb.sheetnames:
        wb.create_sheet('Summary')
    ws = wb['Summary']

    summary_data = final_output["summary"]
    company_name = summary_data["company_name"]
    exchange = summary_data["exchange"]
    symbol = summary_data["symbol"]
    description = summary_data["description"]

    # Write and format the combined title in E1
    combined_title = f"{company_name.upper()} ({exchange}) - {symbol}"
    ws['E1'] = combined_title
    ws['E1'].font = title_font

    # Write "Company Description" in A4 with label formatting
    ws['A4'] = "Company Description"
    ws['A4'].font = label_font

    # Wrap and write the description
    wrapped_lines = textwrap.wrap(description, width=150)
    start_row = 5
    col = 2  # Column B
    for i, line in enumerate(wrapped_lines):
        cell = ws.cell(row=start_row + i, column=col, value=line)
        # Optionally, you can set font here if needed

def write_company_description(writer, final_output):
    reported_currency = final_output["summary"]["reported_currency"]
    cd_info = final_output["company_description"]
    cd_data = cd_info["data"]
    wb = writer.book

    if "Co. Desc" not in wb.sheetnames:
        wb.create_sheet("Co. Desc")
    ws = wb["Co. Desc"]
    ws.freeze_panes = "B1"

    # Write and format labels (unchanged)
    ws.cell(row=1, column=3, value="Currency").fill = label_fill
    ws.cell(row=1, column=3).font = label_font
    ws.cell(row=1, column=3).border = thin_border

    ws.cell(row=1, column=5, value="FY End").fill = label_fill
    ws.cell(row=1, column=5).font = label_font
    ws.cell(row=1, column=5).border = thin_border

    ws.cell(row=1, column=7, value="Stock Price").fill = label_fill
    ws.cell(row=1, column=7).font = label_font
    ws.cell(row=1, column=7).border = thin_border

    ws.cell(row=1, column=9, value="Market Cap").fill = label_fill
    ws.cell(row=1, column=9).font = label_font
    ws.cell(row=1, column=9).border = thin_border

    # Write and format data cells (unchanged)
    rc_cell = ws.cell(row=2, column=3, value=reported_currency)
    rc_cell.fill = data_fill
    rc_cell.font = label_font
    rc_cell.border = thin_border

    fiscal_year_end = cd_info.get("fiscal_year_end")
    fy_cell = ws.cell(row=2, column=5, value=fiscal_year_end)
    fy_cell.fill = data_fill
    fy_cell.font = label_font
    fy_cell.border = thin_border

    stock_price = to_float(cd_info.get("stock_price"))
    sp_cell = ws.cell(row=2, column=7, value=stock_price)
    sp_cell.fill = data_fill
    sp_cell.font = label_font
    sp_cell.number_format = '#,##0.00'
    sp_cell.border = thin_border

    market_cap = to_float(cd_info.get("marketCapitalization")) / 1_000_000
    mc_cell = ws.cell(row=2, column=9, value=market_cap)
    mc_cell.fill = data_fill
    mc_cell.font = label_font
    mc_cell.number_format = '#,##0'
    mc_cell.border = thin_border

    # Sort years for consistent column ordering
    sorted_years = sorted(cd_data.keys(), key=lambda x: int(x))

    # Determine the next two years
    if sorted_years:
        max_year = max(int(year) for year in sorted_years)
        new_years = [str(max_year + 1), str(max_year + 2)]
    else:
        new_years = ["2024", "2025"]

    all_years = sorted_years + new_years

    # Write the years at row=3 starting at column B
    start_col = 2  # Column B
    for i, year in enumerate(all_years):
        year_cell = ws.cell(row=3, column=start_col + i, value=year)
        year_cell.fill = label_fill
        year_cell.font = label_font
        year_cell.border = thin_border

    # Define metrics that should be displayed in millions
    million_scale_metrics = {
        "net_profit", "dividends_paid", "shares_outstanding",
        "buyback", "share_equity", "assets"
    }

    # Define the metrics and their corresponding row positions
    metric_positions = {
        "net_profit": 4, "diluted_eps": 5, "operating_eps": 6, "pe_ratio": 8,
        "price_low": 9, "price_high": 10, "dividends_paid": 12, "dividends_per_share": 13,
        "avg_dividend_yield": 14, "shares_outstanding": 16, "buyback": 17, "share_equity": 19,
        "book_value_per_share": 20, "assets": 22, "return_on_equity": 24, "return_on_assets": 25,
        "leverage_ratio": 26
    }

    # Define human-readable labels for each metric
    metric_labels = {
        "net_profit": "Net Profit", "diluted_eps": "Diluted EPS", "operating_eps": "Operating EPS",
        "pe_ratio": "P/E Ratio", "price_low": "Yrly Price Low", "price_high": "Yrly Price High",
        "dividends_paid": "Dividends Paid", "dividends_per_share": "Dividends/Share",
        "avg_dividend_yield": "Avg Div Yield", "shares_outstanding": "Shares Outstanding",
        "buyback": "Buyback", "share_equity": "Share Equity", "book_value_per_share": "Book Value/Share",
        "assets": "Total Assets", "return_on_equity": "Return on Equity", "return_on_assets": "Return on Assets",
        "leverage_ratio": "Leverage Ratio"
    }

    # Number formats for each metric
    number_formats = {
        "net_profit": '#,##0', "diluted_eps": '#,##0.00', "operating_eps": '#,##0.00',
        "pe_ratio": '#,##0.0', "price_low": '#,##0.00', "price_high": '#,##0.00',
        "dividends_paid": '#,##0', "dividends_per_share": '#,##0.00', "avg_dividend_yield": '0.00%',
        "shares_outstanding": '#,##0', "buyback": '#,##0', "share_equity": '#,##0',
        "book_value_per_share": '#,##0.00', "assets": '#,##0', "return_on_equity": '0.00%',
        "return_on_assets": '0.00%', "leverage_ratio": '#,##0.0'
    }

    # Write metric labels in column A
    for metric, metric_row in metric_positions.items():
        label_cell = ws.cell(row=metric_row, column=1, value=metric_labels[metric])
        label_cell.fill = label_fill
        label_cell.font = label_font
        label_cell.border = thin_border

    # === REVISED FORMULA GENERATION LOGIC ===
    # For each metric, write the data for each year, using formulas where possible
    for metric, metric_row in metric_positions.items():
        for i, year in enumerate(all_years):
            col = start_col + i
            col_letter = get_column_letter(col)
            
            use_formula = True
            formula = ""
            value = None # Used only when use_formula is False

            if metric == "operating_eps":
                if year in new_years:
                    # For new years, calculate from net_profit / shares_outstanding
                    formula = f"={col_letter}{metric_positions['net_profit']}/{col_letter}{metric_positions['shares_outstanding']}"
                else:
                    # For historical years, use the raw data
                    use_formula = False
            elif metric == "diluted_eps":
                formula = f"={col_letter}{metric_positions['net_profit']}/{col_letter}{metric_positions['shares_outstanding']}"
            elif metric == "pe_ratio":
                formula = f"=(({col_letter}{metric_positions['price_low']}+{col_letter}{metric_positions['price_high']})/2)/{col_letter}{metric_positions['diluted_eps']}"
            elif metric == "buyback":
                if i > 0: # Skip first year
                    prev_col_letter = get_column_letter(col - 1)
                    formula = f"=({prev_col_letter}{metric_positions['shares_outstanding']}-{col_letter}{metric_positions['shares_outstanding']})*(({col_letter}{metric_positions['price_low']}+{col_letter}{metric_positions['price_high']})/2)"
                else: # For the first year, no formula is possible
                    use_formula = False
                    value = "N/A"
            elif metric == "dividends_per_share":
                formula = f"={col_letter}{metric_positions['dividends_paid']}/{col_letter}{metric_positions['shares_outstanding']}"
            elif metric == "avg_dividend_yield":
                formula = f"={col_letter}{metric_positions['dividends_per_share']}/((({col_letter}{metric_positions['price_low']}+{col_letter}{metric_positions['price_high']})/2))"
            elif metric == "book_value_per_share":
                formula = f"={col_letter}{metric_positions['share_equity']}/{col_letter}{metric_positions['shares_outstanding']}"
            elif metric == "return_on_equity":
                formula = f"={col_letter}{metric_positions['net_profit']}/{col_letter}{metric_positions['share_equity']}"
            elif metric == "return_on_assets":
                formula = f"={col_letter}{metric_positions['net_profit']}/{col_letter}{metric_positions['assets']}"
            elif metric == "leverage_ratio":
                formula = f"={col_letter}{metric_positions['assets']}/{col_letter}{metric_positions['share_equity']}"
            else:
                # This metric is a primary input, not a calculation
                use_formula = False

            # --- Apply either the formula or the static value ---
            if use_formula:
                data_cell = ws.cell(row=metric_row, column=col, value=formula)
            else:
                if value != "N/A": # If not already set (e.g., for first year buyback)
                    value = cd_data.get(year, {}).get(metric)
                    if value is not None and metric in million_scale_metrics:
                        value /= 1_000_000
                data_cell = ws.cell(row=metric_row, column=col, value=value)

            # Apply common formatting to all data cells
            data_cell.fill = data_fill
            data_cell.border = thin_border
            data_cell.font = data_tnr_italic_font if year in new_years else data_tnr_font
            data_cell.alignment = right_alignment
            if metric in number_formats:
                data_cell.number_format = number_formats[metric]

def write_analyses_sheet(writer, final_output):
    """
    Writes the Analyses sheet with dynamic formulas for all calculated metrics,
    including investment characteristics and the main data table for insurance-specific values.
    """
    reported_currency = final_output["summary"]["reported_currency"]
    analyses = final_output["analyses"]
    # inv_char is no longer needed as we will calculate these values with formulas
    data = analyses["data"]
    wb = writer.book

    if "Analyses" not in wb.sheetnames:
        wb.create_sheet("Analyses")
    ws = wb["Analyses"]
    ws.freeze_panes = "B1"

    # Write and format the "Investment Characteristics" title (unchanged)
    ic_cell = ws.cell(row=1, column=6, value="Investment Characteristics (in mlns " + reported_currency + ")")
    ic_cell.fill = label_fill
    ic_cell.font = title_font
    apply_table_border(ws, 1, 5, 9)
    title_fill_range(ws, 1, 5, 9)

    # Write labels for Investment Characteristics (unchanged)
    labels_with_positions = [
        (3, 3, "Earnings Analysis:"), (5, 4, "Growth Rate %:"), (7, 4, "Quality %:"),
        (3, 8, "Use Of Earnings Analysis:"), (5, 9, "Avg Div Payout Rate:"), (7, 9, "Avg Stk Buyback Rate:"),
        (16, 3, "Premium Analysis:"), (18, 4, "Growth Rate %:"), (20, 4, "Growth Rate PS %:"),
        (16, 8, "Premium Analysis (last 5 yrs.):"), (18, 9, "Growth Rate %:"), (20, 9, "Growth Rate PS %:")
    ]
    for row, col, text in labels_with_positions:
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = label_fill
        cell.font = label_font
        apply_table_border(ws, row, col, col + 1)
        title_fill_range(ws, row, col, col + 1)

    # --- FORMULA GENERATION FOR INVESTMENT CHARACTERISTICS ---
    
    # Get historical year range from Co. Desc sheet to build formulas
    cd_data = final_output["company_description"]["data"]
    sorted_years_hist = sorted(cd_data.keys(), key=lambda x: int(x))
    
    if not sorted_years_hist:
        print("Warning: No historical data found in 'company_description' to build analysis formulas.")
        return

    # Define column letters for the full historical range
    first_year_col = 2
    last_year_col = first_year_col + len(sorted_years_hist) - 1
    first_year_letter = get_column_letter(first_year_col)
    last_year_letter = get_column_letter(last_year_col)
    
    # Define column letters for the last 5 years
    last_5y_count = min(5, len(sorted_years_hist))
    first_5y_col = last_year_col - last_5y_count + 1
    first_5y_letter = get_column_letter(first_5y_col)

    # Define the data cells with formulas instead of static values
    data_cells_formulas = {
        # Earnings Growth Rate (All Years): Using RATE on Operating EPS from Co. Desc (Row 6)
        (5, 5): f"=RATE(COUNT('Co. Desc'!{first_year_letter}6:{last_year_letter}6),,'Co. Desc'!{first_year_letter}6*-1,'Co. Desc'!{last_year_letter}6)",
        
        # Quality % and Payout Rate (Unchanged from previous version)
        (7, 5): f"=AVERAGE('Co. Desc'!{first_year_letter}5:{last_year_letter}5)/AVERAGE('Co. Desc'!{first_year_letter}6:{last_year_letter}6)",
        (5, 10): f"=AVERAGE('Co. Desc'!{first_year_letter}13:{last_year_letter}13)/AVERAGE('Co. Desc'!{first_year_letter}6:{last_year_letter}6)",
        (7, 10): f"=SUM('Co. Desc'!{first_year_letter}17:{last_year_letter}17)/SUM('Co. Desc'!{first_year_letter}4:{last_year_letter}4)",
        
        # Premium Growth Rate (All Years): Using RATE on Premium Earned from this sheet (Row 10)
        (18, 5): f"=RATE(COUNT({first_year_letter}10:{last_year_letter}10),,{first_year_letter}10*-1,{last_year_letter}10)",
        
        # Premium per Share Growth (All Years): Standard CAGR formula is more suitable here
        (20, 5): f"=((({last_year_letter}10/'Co. Desc'!{last_year_letter}16)/({first_year_letter}10/'Co. Desc'!{first_year_letter}16))^(1/({len(sorted_years_hist)}-1)))-1",
        
        # Premium Growth Rate (Last 5 Years): Using RATE on the last 5 years of data
        (18, 10): f"=RATE(COUNT({first_5y_letter}10:{last_year_letter}10),,{first_5y_letter}10*-1,{last_year_letter}10)",
        
        # Premium per Share Growth (Last 5 Years): Standard CAGR on the last 5 years
        (20, 10): f"=((({last_year_letter}10/'Co. Desc'!{last_year_letter}16)/({first_5y_letter}10/'Co. Desc'!{first_5y_letter}16))^(1/({last_5y_count}-1)))-1",
    }

    # Write the investment characteristics data using the new formulas
    for (row, col), formula in data_cells_formulas.items():
        cell = ws.cell(row=row, column=col, value=formula)
        cell.fill = data_fill
        cell.font = data_tnr_bold_font
        cell.border = thin_border
        cell.number_format = '0.0%'

    # --- MAIN DATA TABLE WITH FORMULAS ---

    # Get all years (historical + new)
    max_year = int(sorted_years_hist[-1])
    new_years = [str(max_year + 1), str(max_year + 2)]
    all_years = sorted_years_hist + new_years
    start_col = 2  # Column B

    # Write year headers for the data tables
    for i, year in enumerate(all_years):
        for row in [9, 22]: # Rows for year headers
            cell = ws.cell(row=row, column=start_col + i, value=year)
            cell.fill = label_fill
            cell.font = label_font
            cell.border = thin_border

    # Define metric rows and labels (unchanged)
    metric_rows_1 = { "premium_earned": 10, "benefit_claims": 11, "gross_underwriting_profit": 12, "underwriting_yield_on_asset": 13, "investment_income": 14, "investment_yield_on_asset": 15 }
    metric_rows_2 = { "non_claim_expenses": 23, "expense_yield_on_asset": 24, "tax_rate": 25, "premium_equity_ratio": 26 }
    additional_labels = { 10: "Premium Earned", 11: "Benefit Claims", 12: "Gross Underwriting Profit", 13: "Underwriting Yield on Asset", 14: "Investment Income", 15: "Investment Yield on Asset", 23: "Non-Claim Expenses", 24: "Expense Yield on Asset", 25: "Tax Rate", 26: "Premium/Equity Ratio" }
    number_formats = { "premium_earned": '#,##0', "benefit_claims": '#,##0', "gross_underwriting_profit": '#,##0', "underwriting_yield_on_asset": '0.00%', "investment_income": '#,##0', "investment_yield_on_asset": '0.00%', "non_claim_expenses": '#,##0', "expense_yield_on_asset": '0.00%', "tax_rate": '0.00%', "premium_equity_ratio": '0.00' }
    million_scale_metrics = { "premium_earned", "benefit_claims", "gross_underwriting_profit", "investment_income", "non_claim_expenses" }

    # Write labels for the main data table
    for row, label in additional_labels.items():
        cell = ws.cell(row=row, column=1, value=label)
        cell.fill = label_fill
        cell.font = label_font
        cell.border = thin_border

    # Write the main data table using a mix of static values (historical) and formulas
    all_metric_rows = {**metric_rows_1, **metric_rows_2}
    for metric, row_num in all_metric_rows.items():
        for i, year in enumerate(all_years):
            col = start_col + i
            col_letter = get_column_letter(col)
            
            # Default to no value
            cell_value = None
            is_formula = False

            # --- DEFINE FORMULAS FOR CALCULATED METRICS ---
            if metric == "gross_underwriting_profit":
                cell_value = f"={col_letter}{metric_rows_1['premium_earned']}-{col_letter}{metric_rows_1['benefit_claims']}"
                is_formula = True
            elif metric == "underwriting_yield_on_asset":
                cell_value = f"={col_letter}{metric_rows_1['gross_underwriting_profit']}/'Co. Desc'!{col_letter}22" # Assets are in row 22
                is_formula = True
            elif metric == "investment_yield_on_asset":
                cell_value = f"={col_letter}{metric_rows_1['investment_income']}/'Co. Desc'!{col_letter}22"
                is_formula = True
            elif metric == "expense_yield_on_asset":
                cell_value = f"={col_letter}{metric_rows_2['non_claim_expenses']}/'Co. Desc'!{col_letter}22"
                is_formula = True
            elif metric == "premium_equity_ratio":
                cell_value = f"={col_letter}{metric_rows_1['premium_earned']}/'Co. Desc'!{col_letter}19" # Share Equity is in row 19
                is_formula = True
            
            # For primary data, get value from JSON for historical years
            if not is_formula and year in sorted_years_hist:
                cell_value = data.get(year, {}).get(metric)
                if cell_value is not None and metric in million_scale_metrics:
                    cell_value /= 1_000_000
            
            # Write value to cell
            cell = ws.cell(row=row_num, column=col, value=cell_value)
            cell.fill = data_fill
            cell.font = data_tnr_italic_font if year in new_years else data_tnr_font
            cell.border = thin_border
            cell.number_format = number_formats[metric]
            cell.alignment = right_alignment

def write_profit_desc_sheet(writer, final_output):
    """
    Writes the profit description sheet for balance sheet companies using formulas for calculated fields.

    This revised version incorporates specific logic for sourcing Gross Revenues from the JSON 'total'
    while using the sum of its breakdown for underwriting calculations. It also adds CAGR values for
    all breakdown items and refines cell styling to match the non-BS report aesthetics.
    """
    # 1. SETUP: Get data and create the worksheet
    reported_currency = final_output["summary"]["reported_currency"]
    pd_info = final_output["profit_description"]
    pchar = pd_info["profit_description_characteristics"]
    pdata = pd_info["data"]
    wb = writer.book

    if "Profit.Desc." not in wb.sheetnames:
        wb.create_sheet("Profit.Desc.")
    ws = wb["Profit.Desc."]
    ws.freeze_panes = "D1"

    # Define a clear (white) fill for breakdown items
    no_fill = PatternFill(fill_type=None)

    # Write and format the title
    title_text = f"Description & Analysis of Profitability (in mlns {reported_currency})"
    title_cell = ws.cell(row=1, column=4, value=title_text)
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_fill_range(ws, 1, 3, 10)
    apply_table_border(ws, 1, 3, 10)

    # 2. DEFINE LAYOUT AND DYNAMICALLY MAP ROWS
    metrics_order = [
        "gross_revenues", "investment_income", "internal_costs", "operating_margin",
        "external_costs", "earnings", "equity_employed", "shares_repurchased"
    ]
    metric_labels = {
        "gross_revenues": "Gross Revenues:", "investment_income": "Investment Income:",
        "internal_costs": "Internal Costs:", "operating_margin": "Operating Margin:",
        "external_costs": "External Costs:", "earnings": "Earnings:",
        "equity_employed": "Equity Employed:", "shares_repurchased": "Shares Repurchased:"
    }
    
    all_breakdown_keys = {}
    metrics_with_breakdowns = ["gross_revenues", "internal_costs", "operating_margin", "external_costs"]
    for metric in metrics_with_breakdowns:
        keys = set()
        for year_data in pdata.values():
            mdata = year_data.get(metric, {})
            if isinstance(mdata, dict) and "breakdown" in mdata:
                keys.update(mdata.get("breakdown", {}).keys())
        all_breakdown_keys[metric] = sorted(list(keys))

    # 3. WRITE LABELS AND CAGR VALUES
    current_row = 5
    metric_rows = {}
    breakdown_rows = {}

    for metric in metrics_order:
        # Write main metric label
        label = metric_labels.get(metric)
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col, value=label if col == 1 else None)
            cell.fill = label_fill
            cell.font = label_font
        metric_rows[metric] = current_row
        apply_table_border(ws, current_row, 1, 3)

        # Write main metric CAGR
        cagr_key = f"cagr_{metric}_percent"
        cagr_value = pchar.get(cagr_key)
        if cagr_value is not None:
            cagr_cell = ws.cell(row=current_row, column=3, value=cagr_value)
            cagr_cell.number_format = '0.0%'
            cagr_cell.font = Font(name="Arial", italic=True, size=8)

        current_row += 1

        # Write breakdown labels and their CAGRs
        if metric in all_breakdown_keys:
            for bkey in all_breakdown_keys[metric]:
                b_cell = ws.cell(row=current_row, column=2, value=bkey)
                b_cell.font = Font(italic=True)
                breakdown_rows[(metric, bkey)] = current_row

                # Write breakdown CAGR value
                cagr_dict_key = f"cagr_{metric}_breakdown_percent"
                cagr_item_key = f"cagr_{metric}_{bkey}_percent"
                b_cagr_value = pchar.get(cagr_dict_key, {}).get(cagr_item_key)
                if b_cagr_value is not None:
                    b_cagr_cell = ws.cell(row=current_row, column=3, value=b_cagr_value)
                    b_cagr_cell.number_format = '0.0%'
                    b_cagr_cell.font = Font(name="Arial", italic=True, size=8)
                
                current_row += 1

    # 4. WRITE DATA AND FORMULAS
    sorted_years = sorted(pdata.keys(), key=lambda x: int(x))
    start_col_for_years = 4

    # Write Year Headers
    for i, year in enumerate(sorted_years):
        year_col = start_col_for_years + i * 2
        y_cell = ws.cell(row=3, column=year_col, value=year)
        y_cell.fill = label_fill
        y_cell.font = label_font
        y_cell.border = thin_border
        y_cell.alignment = center_alignment
        if pdata.get(year, {}).get("filing_url"):
            y_cell.hyperlink = pdata[year]["filing_url"]
            y_cell.font = Font(name="Times New Roman", size=10, bold=True, italic=True, underline="single", color="0000FF")

    # Loop through years to populate data
    for i, year in enumerate(sorted_years):
        year_col = start_col_for_years + i * 2
        col_letter = get_column_letter(year_col)
        co_desc_col_letter = get_column_letter(2 + i)
        
        # --- A. Write Primary Data (from JSON) with specific styling ---
        # Gross Revenues (Total) - from JSON 'total'
        gr_total_val = pdata.get(year, {}).get("gross_revenues", {}).get("total")
        ws.cell(row=metric_rows["gross_revenues"], column=year_col, value=to_float(gr_total_val) / 1_000_000 if to_float(gr_total_val) is not None else None)

        # Gross Revenues (Breakdown) - from JSON 'breakdown', NO FILL, NO BORDER
        for bkey in all_breakdown_keys["gross_revenues"]:
            val = pdata.get(year, {}).get("gross_revenues", {}).get("breakdown", {}).get(bkey)
            cell = ws.cell(row=breakdown_rows[("gross_revenues", bkey)], column=year_col, value=to_float(val) / 1_000_000 if to_float(val) is not None else None)
            cell.fill = no_fill
            cell.border = Border()
            cell.font = data_arial_italic_font

        # Investment Income - from JSON
        inv_inc_val = pdata.get(year, {}).get("investment_income")
        ws.cell(row=metric_rows["investment_income"], column=year_col, value=to_float(inv_inc_val) / 1_000_000 if to_float(inv_inc_val) is not None else None)

        # Internal & External Costs (Breakdown) - from JSON, NO FILL, NO BORDER
        for metric_name in ["internal_costs", "external_costs"]:
            for bkey in all_breakdown_keys[metric_name]:
                val = pdata.get(year, {}).get(metric_name, {}).get("breakdown", {}).get(bkey)
                cell = ws.cell(row=breakdown_rows[(metric_name, bkey)], column=year_col, value=to_float(val) / 1_000_000 if to_float(val) is not None else None)
                cell.fill = no_fill
                cell.border = Border()
                cell.font = data_arial_italic_font

        # --- B. Write Formula-Driven Cells ---
        # Helper to get SUM range for a breakdown metric
        def get_sum_range(metric_name):
            bdown_rows = [breakdown_rows.get((metric_name, k)) for k in all_breakdown_keys[metric_name]]
            return ",".join([f"{col_letter}{r}" for r in bdown_rows if r]) if bdown_rows else ""

        # Internal Costs (Total) = Sum of its breakdown
        ws.cell(row=metric_rows["internal_costs"], column=year_col, value=f"=SUM({get_sum_range('internal_costs')})")
        
        # External Costs (Total) = Sum of its breakdown
        ws.cell(row=metric_rows["external_costs"], column=year_col, value=f"=SUM({get_sum_range('external_costs')})")
        
        # Operating Margin (Total) = Gross Revenues (JSON total) + Investment Income - Internal Costs (formula total)
        rev_ref = f"{col_letter}{metric_rows['gross_revenues']}"
        inv_ref = f"{col_letter}{metric_rows['investment_income']}"
        ic_ref = f"{col_letter}{metric_rows['internal_costs']}"
        ws.cell(row=metric_rows["operating_margin"], column=year_col, value=f"={rev_ref}+{inv_ref}-{ic_ref}")

        # Earnings = Operating Margin - External Costs
        om_ref = f"{col_letter}{metric_rows['operating_margin']}"
        ec_ref = f"{col_letter}{metric_rows['external_costs']}"
        ws.cell(row=metric_rows["earnings"], column=year_col, value=f"={om_ref}-{ec_ref}")

        # Equity Employed & Shares Repurchased (from Co. Desc)
        ws.cell(row=metric_rows["equity_employed"], column=year_col, value=f"='Co. Desc'!{co_desc_col_letter}19")
        ws.cell(row=metric_rows["shares_repurchased"], column=year_col, value=f"='Co. Desc'!{co_desc_col_letter}17")

        # --- C. Operating Margin Breakdown Formulas (Special Cases) ---
        share_equity_ref = f"'Co. Desc'!{co_desc_col_letter}19"
        rev_breakdown_sum_ref = f"SUM({get_sum_range('gross_revenues')})"
        
        # Underwriting Margin = SUM(Revenue Breakdown) - Internal Costs
        underwriting_row = breakdown_rows.get(("operating_margin", "underwriting"))
        if underwriting_row:
             ws.cell(row=underwriting_row, column=year_col, value=f"={rev_breakdown_sum_ref}-{ic_ref}")
        
        # Pre-tax Combined Ratio = (SUM(Revenue Breakdown) - Underwriting) / SUM(Revenue Breakdown)
        ptcr_row = breakdown_rows.get(("operating_margin", "pretax_combined_ratio"))
        if ptcr_row and underwriting_row:
            underwriting_ref = f"{col_letter}{underwriting_row}"
            ws.cell(row=ptcr_row, column=year_col, value=f"=IFERROR(({rev_breakdown_sum_ref}-{underwriting_ref})/{rev_breakdown_sum_ref},\"\")")

        # Pre-tax Insurance Yield on Equity = Underwriting / Share Equity
        ptiyoe_row = breakdown_rows.get(("operating_margin", "pretax_insurance_yield_on_equity"))
        if ptiyoe_row and underwriting_row:
            underwriting_ref = f"{col_letter}{underwriting_row}"
            ws.cell(row=ptiyoe_row, column=year_col, value=f"=IFERROR({underwriting_ref}/{share_equity_ref},\"\")")

        # Pre-tax Return on Equity = Operating Margin (Total) / Share Equity
        ptroe_row = breakdown_rows.get(("operating_margin", "pretax_return_on_equity"))
        if ptroe_row:
             ws.cell(row=ptroe_row, column=year_col, value=f"=IFERROR({om_ref}/{share_equity_ref},\"\")")

    # 5. APPLY FINAL FORMATTING & PERCENTAGES
    for r in range(5, current_row):
        is_breakdown_data_row = any(br == r for br in breakdown_rows.values())
        
        for i, year in enumerate(sorted_years):
            year_col = start_col_for_years + i * 2
            cell = ws.cell(row=r, column=year_col)

            # Apply fill and border to main metric rows (not breakdown data rows)
            if not is_breakdown_data_row:
                cell.fill = data_fill
                cell.border = thin_border
            
            # Set font for formula cells
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                 cell.font = data_arial_italic_font

            # Apply number formats
            if cell.data_type == 'f' or isinstance(cell.value, (int, float)): # Formula or Number
                 is_pct_metric = False
                 for (m, bk), br in breakdown_rows.items():
                     if br == r and m == "operating_margin" and bk in ["pretax_combined_ratio", "pretax_insurance_yield_on_equity", "pretax_return_on_equity"]:
                         is_pct_metric = True
                         break
                 cell.number_format = '0.00%' if is_pct_metric else '#,##0'

    # Calculate Percentages of Gross Revenue
    rev_row_num = metric_rows.get("gross_revenues")
    if rev_row_num:
        for i, year in enumerate(sorted_years):
            year_col = start_col_for_years + i * 2
            col_letter = get_column_letter(year_col)
            rev_cell_ref = f"{col_letter}{rev_row_num}"
            
            for r in range(rev_row_num + 1, current_row):
                # Skip rows that are already formatted as percentages
                is_pct_metric = False
                for (m, bk), br in breakdown_rows.items():
                    if br == r and m == "operating_margin" and bk in ["pretax_combined_ratio", "pretax_insurance_yield_on_equity", "pretax_return_on_equity"]:
                        is_pct_metric = True
                        break
                if not is_pct_metric:
                    metric_cell_ref = f"{col_letter}{r}"
                    formula = f"=IFERROR({metric_cell_ref}/{rev_cell_ref},\"\")"
                    cell = ws.cell(row=r, column=year_col + 1, value=formula)
                    cell.font = Font(name="Arial", italic=True, size=8)
                    cell.number_format = '0.0%'

def write_balance_sheet_sheet(writer, final_output):
    """
    Writes a Balance Sheet worksheet to the Excel workbook using data from final_output.
    This version is flexible with the input data structure and preserves the order of breakdown items:
      - If a section (assets, liabilities, shareholders_equity) has a key equal to its name,
        that value is treated as the total for that section.
      - Any other keys in the section (or, if provided, in a nested "breakdown" dict)
        are treated as breakdown items.
    Values are converted from raw units to millions.
    """
    # Assume these style objects and helper functions are defined elsewhere:
    #   label_fill, label_font, title_font, data_fill, data_arial_bold_font,
    #   data_arial_italic_font, thin_border, title_fill_range, apply_table_border, to_float

    reported_currency = final_output["summary"]["reported_currency"]
    bs_info = final_output["balance_sheet"]
    bs_char = bs_info["balance_sheet_characteristics"]
    bs_data = bs_info["data"]
    wb = writer.book

    # Create or get the worksheet
    sheet_name = "Balance Sht."
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    ws.freeze_panes = "F1"

    # Write and format the title
    title_cell = ws.cell(row=1, column=4, value=f"Balance Sheet (in mlns {reported_currency}):")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_fill_range(ws, 1, 4, 7)
    apply_table_border(ws, 1, 4, 7)

    # Extract CAGR values (if available)
    cagr_assets = bs_char.get("cagr_total_assets_percent")
    cagr_liabilities = bs_char.get("cagr_total_liabilities_percent")
    cagr_equity = bs_char.get("cagr_total_shareholders_equity_percent")

    # Define the top-level sections and their labels
    top_sections = [
        ("assets", "Assets:"),
        ("liabilities", "Liabilities:"),
        ("shareholders_equity", "Shareholder's Equity:")
    ]

    # Get the years in sorted order (assumes year keys are numeric strings)
    sorted_years = sorted(bs_data.keys(), key=lambda x: int(x))
    start_col_for_years = 6  # Column F

    # Write and format the header row with years (row 3)
    for i, year in enumerate(sorted_years):
        year_col = start_col_for_years + i
        y_cell = ws.cell(row=3, column=year_col, value=year)
        y_cell.fill = label_fill
        y_cell.font = label_font
        y_cell.border = thin_border

    # Start writing section rows from row 5
    current_row = 5

    for section_key, section_label in top_sections:
        # Write the section label in column A with header formatting.
        sec_label_cell = ws.cell(row=current_row, column=1, value=section_label)
        sec_label_cell.fill = label_fill
        sec_label_cell.font = label_font
        title_fill_range(ws, current_row, 1, 5)
        apply_table_border(ws, current_row, 1, 5)
        total_row = current_row

        # -- Write the total row for the section (if present) --
        # In the new format, we expect that if a total is provided it is stored under the key
        # that is the same as the section_key (e.g. "assets" for assets). Otherwise,
        # there is no total row.
        total_found = False
        for year in sorted_years:
            sec_data = bs_data.get(year, {}).get(section_key, {})
            if section_key in sec_data:
                total_found = True
                break

        if total_found:
            for i, year in enumerate(sorted_years):
                sec_data = bs_data.get(year, {}).get(section_key, {})
                # Use the value from the key equal to the section name (if available)
                raw_val = sec_data.get(section_key)
                val = to_float(raw_val) if raw_val is not None else None
                if val is not None:
                    val = val / 1_000_000  # Convert to millions
                d_cell = ws.cell(row=total_row, column=start_col_for_years + i, value=val)
                d_cell.fill = data_fill
                d_cell.font = data_arial_bold_font
                d_cell.border = thin_border
                if isinstance(val, (int, float)):
                    d_cell.number_format = '#,##0'
            # Apply the corresponding CAGR value in column E (if available)
            if section_key == "assets" and cagr_assets is not None:
                ws.cell(row=total_row, column=5, value=cagr_assets).number_format = '0.0%'
            elif section_key == "liabilities" and cagr_liabilities is not None:
                ws.cell(row=total_row, column=5, value=cagr_liabilities).number_format = '0.0%'
            elif section_key == "shareholders_equity" and cagr_equity is not None:
                ws.cell(row=total_row, column=5, value=cagr_equity).number_format = '0.0%'
            current_row += 1  # Advance to the next row after the total row

        # -- Determine breakdown keys while preserving original order --
        # Instead of sorting keys alphabetically, we iterate over the years and add keys
        # in the order they first appear.
        breakdown_keys = []
        for year in sorted_years:
            sec_data = bs_data.get(year, {}).get(section_key, {})
            if "breakdown" in sec_data and isinstance(sec_data["breakdown"], dict):
                for key in sec_data["breakdown"]:
                    if key not in breakdown_keys:
                        breakdown_keys.append(key)
            else:
                for key in sec_data:
                    if key != section_key and key not in breakdown_keys:
                        breakdown_keys.append(key)

        # -- Write breakdown rows --
        for bkey in breakdown_keys:
            # Write the breakdown label in column B (no special fill)
            ws.cell(row=current_row, column=2, value=bkey)
            # Write data for each year for this breakdown item
            for i, year in enumerate(sorted_years):
                sec_data = bs_data.get(year, {}).get(section_key, {})
                # If a nested breakdown exists, use it; otherwise, read directly.
                if "breakdown" in sec_data and isinstance(sec_data["breakdown"], dict):
                    raw_val = sec_data["breakdown"].get(bkey)
                else:
                    raw_val = sec_data.get(bkey)
                val = to_float(raw_val) if raw_val is not None else None
                if val is not None:
                    val = val / 1_000_000  # convert to millions
                bdata_cell = ws.cell(row=current_row, column=start_col_for_years + i, value=val)
                bdata_cell.font = data_arial_italic_font
                if isinstance(val, (int, float)):
                    bdata_cell.number_format = '#,##0'
            current_row += 1

        # Add a blank line before the next section
        current_row += 1

def to_float(val):
    if val is None or val == "":
        return None
    val_str = str(val).replace("%", "").replace(",", "").strip()
    if val_str == "":
        return None
    try:
        return float(val_str)
    except ValueError:
        return None
    
def write_qualities_sheet(writer, final_output):
    """
    Create or update a sheet called 'Qualities' that displays the text from
    final_output['qualities'] with:
    - Numbered items (1-10)
    - Bold headers inline with numbers
    - Wrapped descriptive text starting on the next line
    Each quality is separated by a blank line for readability.
    """
    # nothing to do if no qualities
    if not final_output.get("qualities"):
        return

    wb = writer.book

    # Create sheet if it doesn't exist, otherwise clear it
    if "Qualities" in wb.sheetnames:
        ws = wb["Qualities"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Qualities")

    # Title row
    ws["A1"] = "Core Analysis"
    ws["A1"].font      = Font(name="Times New Roman", size=14, bold=True)
    ws["A1"].fill      = label_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border    = thin_border

    text = final_output["qualities"].strip()

    # Split on any run of 1+ newlines immediately before "digit+."
    qualities = re.split(r'(?:\r?\n)+(?=\d+\.)', text)

    current_row = 3
    col = 1

    for entry in qualities:
        entry = entry.strip()
        if not entry:
            continue

        # First try to grab markdown-bold header (with or without colon inside)
        m = re.match(r'(\d+)\.\s*\*\*(.*?)\*\*\s*[:]?[\s]*(.*)', entry, re.DOTALL)
        if not m:
            # fallback to plain "1. Header: description"
            m = re.match(r'(\d+)\.\s*(.*?):\s*(.*)', entry, re.DOTALL)
        if not m:
            # if it still fails, skip
            continue

        number, header, description = m.groups()
        header = header.rstrip(':').strip()

        # Write the header line
        hdr = ws.cell(row=current_row, column=col)
        hdr.value = f"{number}. {header}:"
        hdr.font  = Font(name="Arial", size=10, bold=True)
        current_row += 1

        # Wrap and write the description
        for line in textwrap.wrap(description.strip(), width=100):
            c = ws.cell(row=current_row, column=col)
            c.value     = line
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(wrapText=True)
            current_row += 1

        # blank line
        current_row += 1

    # final formatting
    ws.column_dimensions[get_column_letter(col)].width = 110
    ws.sheet_view.showGridLines = False

def write_industry_sheet(writer, final_output):
    """
    Write the Industry sheet with operating and market statistics
    
    Parameters:
    writer: ExcelWriter object
    final_output: Dictionary containing the full output data including industry statistics
    """
    # Early return if industry_comparison is None/null
    if not final_output.get("industry_comparison"):
        return
    
    wb = writer.book
    
    # If the sheet doesn't exist yet, create it
    if "Industry" not in wb.sheetnames:
        wb.create_sheet("Industry")
    ws = wb["Industry"]

    industry_data = final_output["industry_comparison"]
    industry_name = final_output["summary"]["industry"]
    # Write and format the "Industry Overview" title
    title_cell = ws.cell(row=1, column=6, value=f"Industry Comparison: {industry_name}")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    apply_table_border(ws, 1, 4, 8)
    title_fill_range(ws, 1, 4, 8)

    # Write Operating Statistics section
    op_stats_cell = ws.cell(row=3, column=2, value="Operating Statistics:")
    op_stats_cell.fill = label_fill
    op_stats_cell.font = label_font
    op_stats_cell.border = thin_border
    
    # Get companies (tickers)
    companies = list(industry_data["operatingStatistics"].keys())
    
    # Define the operating statistics columns and their formats
    op_stats_columns = {
        "Company": (2, None),          # Column B, no special format
        "Debt(yrs.)": (4, '#,##0.0'),  # Column D
        "Sales": (6, '#,##0'),         # Column F
        "ROC": (8, '0.0%'),            # Column H
        "Operating Margin": (10, '0.0%')  # Column J
    }

    # Write operating statistics headers
    row = 5
    for label, (col, _) in op_stats_columns.items():
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = label_fill
        cell.font = label_font
        cell.border = thin_border
        cell.alignment = center_alignment

    # Locate the most recent year column in Profit.Desc. (row 3, every 2 cols)
    pd_col = None
    om_row = None
    if "Profit.Desc." in wb.sheetnames:
        pd_ws = wb["Profit.Desc."]
        c = 4
        year_cols = []
        while c <= pd_ws.max_column and pd_ws.cell(row=3, column=c).value is not None:
            year_cols.append(c)
            c += 2
        if year_cols:
            pd_col = year_cols[-1]
        # Find the row for "Operating Margin:" in Profit.Desc.
        for r in range(1, pd_ws.max_row + 1):
            if pd_ws.cell(row=r, column=1).value == "Operating Margin:":
                om_row = r
                break

    # Write company data
    for idx, company in enumerate(companies):
        row += 1
        # Company name
        ws.cell(row=row, column=2, value=company).font = data_arial_font
        ws.cell(row=row, column=2).alignment = center_alignment
        company_data = industry_data["operatingStatistics"][company]

        for label, (col, fmt) in op_stats_columns.items():
            if label == "Company":
                continue

            # First company’s operating margin comes from Profit.Desc.
            if idx == 0 and label == "Operating Margin" and pd_col and om_row:
                col_letter = get_column_letter(pd_col+1)
                formula = f"='Profit.Desc.'!{col_letter}{om_row}"
                cell = ws.cell(row=row, column=col, value=formula)
                cell.number_format = fmt
            else:
                value = company_data[label]
                if label == "Sales":
                    value = value / 1_000_000
                cell = ws.cell(row=row, column=col, value=value)
                if fmt:
                    cell.number_format = fmt

            cell.font = data_arial_font
            cell.alignment = center_alignment

    last_op_stats_row = row

    # Market Statistics section
    market_stats_start_row = last_op_stats_row + 3
    market_stats_cell = ws.cell(row=market_stats_start_row, column=2, value="Market Statistics:")
    market_stats_cell.fill = label_fill
    market_stats_cell.font = label_font
    market_stats_cell.border = thin_border

    market_stats_columns = {
        "Company": (2, None),
        "P/B":      (4, '#,##0.00'),
        "P/E":      (6, '#,##0.0'),
        "Div. Yld.":(8, '0.00%'),
        "EV/Sales": (10,'#,##0.00'),
    }

    # Headers
    row = market_stats_start_row + 2
    for label, (col, _) in market_stats_columns.items():
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = label_fill
        cell.font = label_font
        cell.border = thin_border
        cell.alignment = center_alignment

    # Data rows
    for idx, company in enumerate(companies):
        row += 1
        # Company name
        ws.cell(row=row, column=2, value=company).font = data_arial_font
        ws.cell(row=row, column=2).alignment = center_alignment
        stats = industry_data["marketStatistics"][company]

        for label, (col, fmt) in market_stats_columns.items():
            if label == "Company":
                continue
            value = stats[label]
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_arial_font
            cell.alignment = center_alignment
            if fmt:
                cell.number_format = fmt

    # Adjust column widths
    for c in range(1, 11):
        ws.column_dimensions[get_column_letter(c)].width = 15

def write_segmentation_sheet(writer, final_output):
    """
    Write the Segmentation sheet showing revenue breakdown by business segment over time.
    
    Parameters:
      writer: ExcelWriter object
      final_output: Dictionary containing the full output data including segmentation data.
    """
    # Early return if segmentation data is not present
    if not final_output.get("segmentation"):
        return

    reported_currency = final_output["summary"]["reported_currency"]
    segmentation_data = final_output["segmentation"]["data"]
    wb = writer.book

    # Create the sheet if it doesn't exist
    if "Segmentation" not in wb.sheetnames:
        wb.create_sheet("Segmentation")
    ws = wb["Segmentation"]
    ws.freeze_panes = "B1"

    # Write and format the title
    title_cell = ws.cell(row=1, column=5, value=f"Revenue Segmentation (in mlns {reported_currency})")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    apply_table_border(ws, 1, 3, 7)
    title_fill_range(ws, 1, 3, 7)

    # Get sorted years and all unique segments from the "breakdown" dictionaries
    sorted_years = sorted(segmentation_data.keys(), key=lambda x: int(x))
    all_segments = set()
    for year in sorted_years:
        # For each year, get the breakdown sub-dictionary
        breakdown = segmentation_data[year].get("breakdown", {})
        all_segments.update(breakdown.keys())
    sorted_segments = sorted(all_segments)

    # Write year headers starting at row 3 (starting at column B)
    for i, year in enumerate(sorted_years):
        year_col = i + 2  # Column B is index 2
        y_cell = ws.cell(row=3, column=year_col, value=year)
        y_cell.fill = label_fill
        y_cell.font = label_font
        y_cell.border = thin_border
        y_cell.alignment = center_alignment

    # Write segment data for each segment (one segment per row)
    current_row = 4
    for segment in sorted_segments:
        # Write the segment name in column A
        segment_cell = ws.cell(row=current_row, column=1, value=segment)
        segment_cell.font = label_font
        segment_cell.fill = label_fill
        segment_cell.border = thin_border

        # Write values for each year
        for i, year in enumerate(sorted_years):
            year_col = i + 2  # Data starts at column B
            # Retrieve the breakdown for the given year and get the segment value
            breakdown = segmentation_data[year].get("breakdown", {})
            value = breakdown.get(segment)

            # Create the cell and apply the common styling
            value_cell = ws.cell(row=current_row, column=year_col)
            value_cell.fill = data_fill
            value_cell.font = data_arial_font
            value_cell.border = thin_border
            value_cell.alignment = right_alignment

            if value is not None:
                # Convert the value to millions and set the cell value/formatting
                value_cell.value = value / 1_000_000
                value_cell.number_format = '#,##0'

        current_row += 1

    # Calculate and write growth rates (CAGR) in the rightmost column.
    # The growth column is placed one column after the last year's column plus an extra spacer.
    growth_col = len(sorted_years) + 3
    growth_header = ws.cell(row=3, column=growth_col, value="CAGR")
    growth_header.fill = label_fill
    growth_header.font = label_font
    growth_header.border = thin_border
    growth_header.alignment = center_alignment

    # For each segment, calculate the compound annual growth rate (CAGR)
    for segment_idx, segment in enumerate(sorted_segments):
        row = segment_idx + 4

        # Get the first valid value for the segment (from the earliest year)
        first_val = None
        for year in sorted_years:
            breakdown = segmentation_data[year].get("breakdown", {})
            val = breakdown.get(segment)
            if val is not None:
                first_val = val
                break

        # Get the last valid value for the segment (from the most recent year)
        last_val = None
        for year in reversed(sorted_years):
            breakdown = segmentation_data[year].get("breakdown", {})
            val = breakdown.get(segment)
            if val is not None:
                last_val = val
                break

        if first_val is not None and last_val is not None and first_val != 0:
            years_between = len(sorted_years) - 1
            if years_between > 0:
                cagr = (last_val / first_val) ** (1 / years_between) - 1
                growth_cell = ws.cell(row=row, column=growth_col, value=cagr)
                growth_cell.number_format = '0.0%'
                growth_cell.font = Font(name="Arial", size=8, italic=True)
                growth_cell.fill = data_fill
                growth_cell.border = thin_border
        else:
            # Create an empty growth cell with consistent styling
            growth_cell = ws.cell(row=row, column=growth_col)
            growth_cell.fill = data_fill
            growth_cell.border = thin_border

    # Adjust column widths: set column A wider for segment names, and other columns to a fixed width.
    ws.column_dimensions['A'].width = 30  # Segment names
    for col in range(2, growth_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

def write_hist_pricing_sheet(writer, final_output):
    """
    Write the Historical Pricing sheet with average ratios and price implications in a 2x2 grid layout
    
    Parameters:
    writer: ExcelWriter object
    final_output: Dictionary containing the full output data including historical pricing and current metrics
    """
    wb = writer.book
    
    if "Hist. Pricing" not in wb.sheetnames:
        wb.create_sheet("Hist. Pricing")
    ws = wb["Hist. Pricing"]

    hist_pricing = final_output.get("historical_pricing", {})
    reported_currency = final_output["summary"]["reported_currency"]

    # Write and format the title
    title_cell = ws.cell(row=1, column=4, value=f"Historical Pricing Analysis ({reported_currency})")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    apply_table_border(ws, 1, 3, 6)
    title_fill_range(ws, 1, 3, 6)

    # Get the first new year column letter from Co. Desc sheet for formulas
    sorted_years = sorted(final_output["company_description"]["data"].keys(), key=lambda x: int(x))
    if sorted_years:
        max_year = max(int(year) for year in sorted_years)
        new_year = str(max_year + 1)
        # In Co. Desc sheet, years start at column B (2)
        first_new_year_col = get_column_letter(2 + len(sorted_years))
    else:
        first_new_year_col = "B"  # fallback

    # Define the grid positions for each metric
    metrics = {
        # Top Left - P/E Ratio
        "P/E Ratio": {
            "low_key": "avg_pe_low",
            "high_key": "avg_pe_high",
            "current_formula": f"='Co. Desc'!{first_new_year_col}5",  # References diluted EPS
            "start_row": 3,
            "start_col": 2,
            "format": '#,##0.0',
            "value_type": "earnings"
        },
        # Top Right - P/Assets Ratio
        "P/Assets Ratio": {
            "low_key": "avg_ps_low",
            "high_key": "avg_ps_high",
            "current_formula": f"='Analyses'!{first_new_year_col}22 / 'Analyses'!{first_new_year_col}16",  # References Sales/Share
            "start_row": 3,
            "start_col": 8,
            "format": '#,##0.00',
            "value_type": "sales"
        },
        # Bottom Left - P/B Ratio
        "P/B Ratio": {
            "low_key": "avg_pb_low",
            "high_key": "avg_pb_high",
            "current_formula": f"='Co. Desc'!{first_new_year_col}20",  # References Book Value/Share
            "start_row": 10,
            "start_col": 2,
            "format": '#,##0.00',
            "value_type": "book_value"
        },
        # Bottom Right - P/CF Ratio
        "P/CF Ratio": {
            "low_key": "avg_pcf_low",
            "high_key": "avg_pcf_high",
            "current_formula": f"=('Co. Desc'!{first_new_year_col}4+'Analyses'!{first_new_year_col}22)/'Co. Desc'!{first_new_year_col}16",  # (Net Profit + Depreciation) / Shares Outstanding
            "start_row": 10,
            "start_col": 8,
            "format": '#,##0.00',
            "value_type": "cash_flow"
        }
    }

    # Write each metric box
    for metric, props in metrics.items():
        start_row = props["start_row"]
        start_col = props["start_col"]
        
        # Write box title
        title_cell = ws.cell(row=start_row, column=start_col, value=metric)
        title_cell.fill = label_fill
        title_cell.font = label_font
        title_cell.alignment = center_alignment
        ws.merge_cells(start_row=start_row, start_column=start_col, 
                      end_row=start_row, end_column=start_col + 1)
        
        # Write metric rows vertically
        metrics_data = [
            ("Used", props["current_formula"]),  # Now using formula instead of value
            ("Avg Low", hist_pricing.get(props["low_key"])),
            ("Avg High", hist_pricing.get(props["high_key"]))
        ]
        
        for idx, (label, value) in enumerate(metrics_data):
            # Write label
            label_cell = ws.cell(row=start_row + 1 + idx, column=start_col, value=label)
            label_cell.fill = label_fill
            label_cell.font = label_font
            label_cell.border = thin_border
            label_cell.alignment = center_alignment
            
            # Write value or formula
            if idx == 0:  # "Used" row
                value_cell = ws.cell(row=start_row + 1 + idx, column=start_col + 1)
                value_cell.value = value  # This is now a formula
            else:
                value_cell = ws.cell(row=start_row + 1 + idx, column=start_col + 1, value=value)
            
            value_cell.fill = data_fill
            value_cell.font = data_arial_font
            value_cell.border = thin_border
            value_cell.number_format = props["format"]
            value_cell.alignment = right_alignment
        
        # Write Buy and Sell rows with formulas
        used_cell = f"{get_column_letter(start_col + 1)}{start_row + 1}"
        avg_low_cell = f"{get_column_letter(start_col + 1)}{start_row + 2}"
        avg_high_cell = f"{get_column_letter(start_col + 1)}{start_row + 3}"
        
        # Write Buy row (Used * Avg Low)
        buy_label = ws.cell(row=start_row + 4, column=start_col, value="Buy")
        buy_label.fill = label_fill
        buy_label.font = label_font
        buy_label.border = thin_border
        buy_label.alignment = center_alignment
        
        buy_cell = ws.cell(row=start_row + 4, column=start_col + 1, 
                          value=f"={used_cell}*{avg_low_cell}")
        buy_cell.fill = data_fill
        buy_cell.font = data_arial_bold_font
        buy_cell.border = thin_border
        buy_cell.number_format = '#,##0.00'
        buy_cell.alignment = right_alignment
        
        # Write Sell row (Used * Avg High)
        sell_label = ws.cell(row=start_row + 5, column=start_col, value="Sell")
        sell_label.fill = label_fill
        sell_label.font = label_font
        sell_label.border = thin_border
        sell_label.alignment = center_alignment
        
        sell_cell = ws.cell(row=start_row + 5, column=start_col + 1, 
                           value=f"={used_cell}*{avg_high_cell}")
        sell_cell.fill = data_fill
        sell_cell.font = data_arial_bold_font
        sell_cell.border = thin_border
        sell_cell.number_format = '#,##0.00'
        sell_cell.alignment = right_alignment
        
        # Add border around the entire box
        for row in range(start_row, start_row + 6):
            for col in range(start_col, start_col + 2):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

    # Adjust column widths
    for base_col in [2, 8]:  # Starting columns for left and right sections
        ws.column_dimensions[get_column_letter(base_col)].width = 12      # Labels
        ws.column_dimensions[get_column_letter(base_col + 1)].width = 20  # Values

def write_valuation_sheet(writer, final_output, ticker):
    """Write the valuation analysis sheet with 2x2 grid layout"""
    reported_currency = final_output["summary"]["reported_currency"]
    wb = writer.book
    
    if "Valuation" not in wb.sheetnames:
        wb.create_sheet("Valuation")
    ws = wb["Valuation"]

    # Title
    title_cell = ws.cell(row=1, column=4, value="Valuation (USD)")
    title_cell.fill = label_fill
    title_cell.font = title_font
    title_fill_range(ws, 1, 3, 5)
    apply_table_border(ws, 1, 3, 5)

    current_price = get_current_quote_yahoo(ticker)
    tbond_rate = get_long_term_rate()
    
    # Get forecast year column reference
    sorted_years = sorted(
        final_output["company_description"]["data"].keys(),
        key=lambda x: int(x)
    )
    first_forecast_col = (
        get_column_letter(2 + len(sorted_years)) if sorted_years else "B"
    )
    last_col_bal_sht = (
        get_column_letter(5 + len(sorted_years)) if sorted_years else "F"
    )
    # =========================================================================
    # REARRANGED SETTINGS
    # =========================================================================
    #   B3 => ADR Multiple
    #   B4 => Currency Ratio
    #   D3 => EPS Growth
    #   D4 => Dividend Growth
    #   F3 => Purchase Discount
    #   F4 => Sell Discount
    #   F5 => PE Multiple
    # =========================================================================
    settings = {
        (3, 2): ("ORD:ADR:", 1),                           # B3
        (4, 2): (f"USD:{reported_currency} rate:", 1),          # B4
        (3, 4): ("EPS growth rate:", 0.10),                     # D3
        (4, 4): ("Dividend growth rate:", 0.10),                # D4
        (3, 6): ("Purchase Discount:", 0.14),                   # F3
        (4, 6): ("Sell Discount:", 0.05),                       # F4
        (5, 6): ("PE Multiple:", 25),                           # F5
    }

    for (row, col), (label, value) in settings.items():
        label_cell = ws.cell(row=row, column=col - 1, value=label)
        label_cell.fill = label_fill
        label_cell.font = label_font
        label_cell.border = thin_border
        
        value_cell = ws.cell(row=row, column=col, value=value)
        value_cell.fill = data_fill
        value_cell.font = data_arial_font
        value_cell.border = thin_border
        value_cell.alignment = center_alignment

        # Format anything < 1 as percentage
        if isinstance(value, float) and value < 1:
            value_cell.number_format = '0.00%'
    
    new_settings = {
        (3, 7): ("Buy %:", 0.60),    # G3 (label), H3 (value = 60%)
        (4, 7): ("Sell %:", 1.20),   # G4 (label), H4 (value = 120%)
    }

    for (row, col), (label, value) in new_settings.items():
        label_cell = ws.cell(row=row, column=col, value=label)
        label_cell.fill = label_fill
        label_cell.font = label_font
        label_cell.border = thin_border
        
        value_cell = ws.cell(row=row, column=col + 1, value=value)
        value_cell.fill = data_fill
        value_cell.font = data_arial_font
        value_cell.border = thin_border
        value_cell.alignment = center_alignment

        # Format them as percentages
        value_cell.number_format = '0.00%'

    # If you need the numeric value of PE multiple in code:
    # pe_multiple_val = ws.cell(row=5, column=6).value  # F5

    # =========================================================================
    # 2x2 GRID LAYOUT
    # =========================================================================
    # We keep top-left and bottom-left in columns B/C.
    # We move top-right and bottom-right to columns E/F.
    # Then fix all formula references accordingly.
    # =========================================================================
    grid_segments = {
        # ---------------------------------------------------------------------
        # TOP LEFT (columns B/C)
        # ---------------------------------------------------------------------
        "Initial Rate of Investment:": {
            "start_row": 8,
            "start_col": 2,  # B
            "metrics": {
                "Current Price:": f"={current_price}",
                # Currency ratio is B4, ADR multiple is B3:
                "Current EPS:": f"='Co. Desc'!{first_forecast_col}5 * B4 * B3",
                # = B10 / B9 once written to the sheet
                "Initial ROI:": "=B10/B9",
            },
        },
        # ---------------------------------------------------------------------
        # TOP RIGHT (columns E/F)
        # ---------------------------------------------------------------------
        "Relative Value to Investment In T-Bonds:": {
            "start_row": 8,
            "start_col": 5,  # E
            "metrics": {
                # Same logic as top-left for 'Current EPS'
                "Current EPS:": f"='Co. Desc'!{first_forecast_col}5 * B4 * B3",
                "T-Bond Rate:": tbond_rate,
                # Was =H9/H10 in original; now =E9/E10
                "Relative Value:": "=E9/E10",
            },
        },
        # ---------------------------------------------------------------------
        # BOTTOM LEFT (columns B/C)
        # ---------------------------------------------------------------------
        "Valuation as an Equity Bond:": {
            "start_row": 15,
            "start_col": 2,  # B
            "metrics": {
                # Use B3 (ADR multiple) and B4 (currency ratio)
                "Current BV:": f"='Co. Desc'!{first_forecast_col}20 * B3 * B4",
                "Current ROE:": f"='Co. Desc'!{first_forecast_col}24",
                "Retained % adjustment:": 0.10,
                # unchanged, presumably references other sheet cells
                "Retained %:": "=1 - 'Analyses'!J5 - 'Analyses'!J7 - B18",
                "Net BV growth:": "=B17*B19",
                "BV in year 10:": "=FV(B20, 10, , -B16)",
                "EPS Adjustment Factor:": 1.5,
                "EPS in Year 10:": "=B17 * B21 * B22",
                # PE multiple is at F5
                "Value at PE Multiple:": "=F5 * B23",
                # Dividend growth is at D4 ⇒ use FV(D4,10,…)
                "Total Dividends:": (
                    f"=(('Co. Desc'!{first_forecast_col}13 + "
                    f"FV(D4, 10, , -'Co. Desc'!{first_forecast_col}13))/2)*10*B3*B4"
                ),
                "Total Future Value:": "=B24+B25",
                # Purchase discount is at F3
                "Purchase at Discount:": "=PV(F3, 10, , B26)*-1",
            },
        },
        # ---------------------------------------------------------------------
        # BOTTOM RIGHT (columns E/F)
        # ---------------------------------------------------------------------
        "Valuation on Earnings Growth:": {
            "start_row": 15,
            "start_col": 5,  # E
            "metrics": {
                # Same B3/B4 for ADR/currency
                "Current EPS:": f"='Co. Desc'!{first_forecast_col}5 * B4 * B3",
                # EPS Growth is at D3 => FV(D3,10,…)
                "EPS in year 10:": "=FV(D3, 10, , -E16)",
                "Avg PE Ratio:": f"=AVERAGE('Co. Desc'!B8:{first_forecast_col}8)",
                # Was =B5*H17 + B25; now =F5*E17 + B25
                "Value at PE Multiple:": "=F5*E17 + B25",
                # Was =RATE(10, , B9, -H19 + B25); now =RATE(10, , B9, -E19 + B25)
                "Price Return:": "=RATE(10, , B9, -E19 + B25)",
                "Dividend Return:": f"='Co. Desc'!{first_forecast_col}14",
                # Was =H20+H21; now =E20+E21
                "Total Return:": "=E20 + E21",
                # Purchase discount is F3
                "Purchase at Discount:": "=PV(F3, 10, , -E19)",
                # Sell discount is F4
                "Sell at Discount:": "=PV(F4, 10, , -E19)",
            },
        },
        "Float Valuation Approach:": {
            "start_row": 8,
            "start_col": 8,  # G
            "metrics": {
                "Float:": f"=('Balance Sht.'!{last_col_bal_sht}24 + 'Balance Sht.'!{last_col_bal_sht}23 -'Balance Sht.'!{last_col_bal_sht}13 -'Balance Sht.'!{last_col_bal_sht}11)*B4",
                
                "Float Growth Rate:": 0.05,
                
                "Cost of Float:": -0.04,
                
                "10yr FV:": "=FV(H10,10,,-H9)",
                
                "Investment Return:": tbond_rate,
                
                "Return on Float:": "=(H13-H11)",
                
                "Tax Burden on Float:": f"='Analyses'!{first_forecast_col}25 * H14",

                "After Tax Return on Float:": "=(H14-H15)",

                "Income on Float:": "=(H12 * H16)",

                "Discount Rate:": 0.08,

                "Capitalization Factor:": "=(H18-H10)",

                "Value of Float EOY10:": "=(H17/H19)",

                "RFR:": 0.06,

                "PV at RFR:": "=PV(H21,10,,-H20)",

                "Value of Insurance Equity:": f"='Co. Desc'!{first_forecast_col}19 * B4",

                "Total Value:": "=(H22+H23)",

                "Shares Outstanding:": f"='Co. Desc'!{first_forecast_col}16 * (1/B3)",
                
                # Share Value = Total Value / Shares Outstanding
                "Per Share Value:": "=(H24 / H25)",
                
                # Buy at => Share Value * Buy % (which is in H1)
                "Buy At:": "=H26 * $H$3",
                
                # Sell at => Share Value * Sell % (which is in H2)
                "Sell At:": "=H26 * $H$4",
            },
        }
    }

    #
    # Write each 2×2 grid segment
    #
    for title, config in grid_segments.items():
        start_row = config["start_row"]
        start_col = config["start_col"]
        
        # Segment title
        title_cell = ws.cell(row=start_row, column=start_col - 1, value=title)
        title_cell.fill = label_fill
        title_cell.font = Font(name="Times New Roman", size=12, bold=True, italic=True)
        title_cell.border = thin_border
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Merge the two columns for the label
        ws.merge_cells(
            start_row=start_row, start_column=start_col - 1, 
            end_row=start_row, end_column=start_col
        )
        
        # Write metrics
        current_row = start_row + 1
        for label, formula in config["metrics"].items():
            # Label
            label_cell = ws.cell(row=current_row, column=start_col - 1, value=label)
            label_cell.fill = label_fill
            label_cell.font = label_font
            label_cell.border = thin_border
            
            # Value/Formula
            value_cell = ws.cell(row=current_row, column=start_col, value=formula)
            value_cell.fill = data_fill
            value_cell.font = data_arial_font
            value_cell.border = thin_border
            value_cell.alignment = center_alignment

            # Make certain items bold
            if label.lower() in ["relative value:", "purchase at discount:", "sell at discount:", "buy at:", "sell at:"]:
                value_cell.font = data_arial_bold_font
            
            # Format numeric cells
            label_lower = label.lower()
            if label_lower == "t-bond rate:":
                value_cell.number_format = '0.00%'
            elif label_lower == "eps adjustment factor:":
                value_cell.number_format = '0.00'
            elif label_lower == "avg pe ratio:":
                value_cell.number_format = '0.00'
            elif label_lower == "rfr:":
                value_cell.number_format = '0.00%'
            elif label_lower == "float:":
                value_cell.number_format = '"$"#,##0.00'
            elif "net bv growth" in label_lower:
                value_cell.number_format = '0.00%'
            elif "shares outstanding" in label_lower:
                value_cell.number_format = '#,##0'
            elif any(x in label_lower for x in ["rate", "tax", "factor", "roi", "return", "roe", "%", "cost"]):
                value_cell.number_format = '0.00%'
            elif any(x in label_lower for x in ["price", "value", "eps", "bv", "dividends", "purchase", "sell",
                                                "ebit", "value", "income", "debt", "share", "pv", "fv", "buy at", "sell at"]):
                value_cell.number_format = '"$"#,##0.00'
            
            current_row += 1

    # =========================================================================
    # Column Widths
    # =========================================================================
    # Example adjustments for label columns vs. value columns
    ws.column_dimensions[get_column_letter(1)].width = 25  # A
    ws.column_dimensions[get_column_letter(2)].width = 15  # B
    ws.column_dimensions[get_column_letter(3)].width = 25  # C
    ws.column_dimensions[get_column_letter(4)].width = 20  # D
    ws.column_dimensions[get_column_letter(5)].width = 25  # E
    ws.column_dimensions[get_column_letter(6)].width = 15  # F
    ws.column_dimensions[get_column_letter(7)].width = 25  # G
    ws.column_dimensions[get_column_letter(8)].width = 15  # H


def generate_config_note(ticker, wb):
    """
    Add a note in cell B1 of the profit_desc sheet if there are any
    configuration overrides for the company from financial_data_config.json.
    
    Args:
        ticker (str): Company ticker symbol
        wb (openpyxl.Workbook): Excel workbook object
    """
    import json
    import logging
    from pathlib import Path

    logger = logging.getLogger(__name__)
    
    # Get the profit_desc sheet
    try:
        sheet = wb["Profit.Desc."]
    except KeyError:
        logger.warning("profit_desc sheet not found in workbook")
        return
    
    # Load the configuration file
    try:
        with open('financial_data_config.json', 'r') as f:
            config = json.load(f)
    except FileNotFoundError:
        logger.warning("financial_data_config.json not found")
        return
    except json.JSONDecodeError:
        logger.warning("Error parsing financial_data_config.json")
        return
    
    # Check if we have any configurations for this ticker
    ticker_config = config.get(ticker.upper(), {})
    if not ticker_config:
        return
    
    # Build the configuration note
    notes = []
    for statement, fields in ticker_config.items():
        field_notes = []
        for field, value in fields.items():
            field_notes.append(f"set {field} to {value}")
        if field_notes:
            notes.append(f"In {statement}, {', '.join(field_notes)}")
    
    if notes:
        note = '. '.join(notes)
        # Add note to cell B1
        sheet['A2'] = f"Configuration overrides: {note}."
        
        # Style the cell
        from openpyxl.styles import Font, PatternFill
        cell = sheet['B1']
        cell.font = Font(italic=True, size=9, color="666666")
        cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

def sync_data_from_profit_desc_bs(writer):
    """
    Updates the Analyses sheet with formula references to the Profit.Desc. sheet
    for key insurance metrics. This ensures data is synced between the two sheets.

    Mappings:
    - Analyses!Premium Earned      <- Profit.Desc!Gross Revenues
    - Analyses!Benefit Claims      <- Profit.Desc!losses_and_expenses
    - Analyses!Investment Income   <- Profit.Desc!Investment Income
    - Analyses!Non-Claim Expenses  <- SUM(Profit.Desc!acquisition_costs + underwriting_expenses)
    """
    wb = writer.book

    # 1. Ensure both required sheets exist
    if "Analyses" not in wb.sheetnames or "Profit.Desc." not in wb.sheetnames:
        print("Warning: Cannot sync data - 'Analyses' or 'Profit.Desc.' sheet not found.")
        return

    analyses_ws = wb["Analyses"]
    profit_desc_ws = wb["Profit.Desc."]

    # 2. Define the static row numbers in the 'Analyses' sheet
    analyses_target_rows = {
        "premium_earned": 10,
        "benefit_claims": 11,
        "investment_income": 14,
        "non_claim_expenses": 23
    }

    # 3. Dynamically find the source row numbers in the 'Profit.Desc.' sheet
    pd_source_rows = {}
    for row in range(1, profit_desc_ws.max_row + 1):
        main_label = profit_desc_ws.cell(row=row, column=1).value
        breakdown_label = profit_desc_ws.cell(row=row, column=2).value
        if main_label:
            pd_source_rows[main_label] = row
        if breakdown_label:
            pd_source_rows[breakdown_label] = row
            
    # Check if all required source labels were found
    required_keys = [
        "Gross Revenues:", "Investment Income:", "losses_and_expenses",
        "acquisition_costs", "underwriting_expenses"
    ]
    if not all(key in pd_source_rows for key in required_keys):
        print("Warning: Could not find all required source rows in Profit.Desc. sheet. Sync aborted.")
        return

    # 4. Get and match the year columns from both sheets
    analyses_years = []
    col = 2  # Analyses years start at column B
    while True:
        year_cell = analyses_ws.cell(row=9, column=col)
        if year_cell.value is None: break
        analyses_years.append((col, str(year_cell.value)))
        col += 1

    profit_desc_years = []
    col = 4  # Profit.Desc. years start at column D
    while True:
        year_cell = profit_desc_ws.cell(row=3, column=col)
        if year_cell.value is None or col > profit_desc_ws.max_column: break
        profit_desc_years.append((col, str(year_cell.value)))
        col += 2  # Skip percentage column

    # 5. Loop through matched years and write the formulas
    for a_col, a_year in analyses_years:
        for pd_col, pd_year in profit_desc_years:
            if a_year == pd_year:
                pd_col_letter = get_column_letter(pd_col)

                # --- Create and write each formula ---

                # Premium Earned <- Gross Revenues
                formula_pe = f"='Profit.Desc.'!{pd_col_letter}{pd_source_rows['Gross Revenues:']}"
                cell_pe = analyses_ws.cell(row=analyses_target_rows['premium_earned'], column=a_col, value=formula_pe)
                cell_pe.number_format = '#,##0'

                # Benefit Claims <- losses_and_expenses
                formula_bc = f"='Profit.Desc.'!{pd_col_letter}{pd_source_rows['losses_and_expenses']}"
                cell_bc = analyses_ws.cell(row=analyses_target_rows['benefit_claims'], column=a_col, value=formula_bc)
                cell_bc.number_format = '#,##0'

                # Investment Income <- Investment Income
                formula_ii = f"='Profit.Desc.'!{pd_col_letter}{pd_source_rows['Investment Income:']}"
                cell_ii = analyses_ws.cell(row=analyses_target_rows['investment_income'], column=a_col, value=formula_ii)
                cell_ii.number_format = '#,##0'
                
                # Non-Claim Expenses <- SUM(acquisition_costs + underwriting_expenses)
                acq_costs_ref = f"'Profit.Desc.'!{pd_col_letter}{pd_source_rows['acquisition_costs']}"
                uw_exp_ref = f"'Profit.Desc.'!{pd_col_letter}{pd_source_rows['underwriting_expenses']}"
                formula_nce = f"=SUM({acq_costs_ref},{uw_exp_ref})"
                cell_nce = analyses_ws.cell(row=analyses_target_rows['non_claim_expenses'], column=a_col, value=formula_nce)
                cell_nce.number_format = '#,##0'

                break # Move to the next year in Analyses sheet

    print("Successfully synced data from Profit.Desc. to Analyses sheet.")

def generate_excel_for_ticker_year(ticker: str, year: int):
    """
    Generate the Excel file for the given ticker and year, writing to:
       ./output/{ticker}.{last 2 digits of year}.2.xlsx
    
    :param ticker: The company symbol/ticker.
    :param year:   The 4-digit year (e.g., 2024). We'll use only the last two digits in the filename.
    """
    # Convert to uppercase and build the filename, e.g.: ./output/ABC.24.2.xlsx
    ticker = ticker.upper()
    year_2_digits = str(year)[-2:]  # last two chars of the year
    xls_filename = os.path.join("output", f"{ticker}.{year_2_digits}.2.xlsx")

    # 1. Load the JSON data
    final_output = load_final_output(ticker)

    # 2. Create or append to the Excel file
    writer = create_xls(xls_filename)

    # 3. Write data to each sheet
    write_summary_sheet(writer, final_output)
    write_company_description(writer, final_output)
    write_analyses_sheet(writer, final_output)
    write_profit_desc_sheet(writer, final_output)
    write_balance_sheet_sheet(writer, final_output)
    write_qualities_sheet(writer, final_output)
    write_industry_sheet(writer, final_output)
    write_segmentation_sheet(writer, final_output)
    write_hist_pricing_sheet(writer, final_output)
    write_valuation_sheet(writer, final_output, ticker)
    generate_config_note(ticker, writer.book)

    sync_data_from_profit_desc_bs(writer)
    # 4. Apply workbook formatting (remove gridlines, etc.)
    format_workbook(writer)

    # 5. Save
    writer.close()
    print(f"Data for {ticker} written to {xls_filename} successfully.")


if __name__ == "__main__":
    # Usage: python write_excel.py TICKER target_file.xls
    if len(sys.argv) < 3:
        print("Usage: python gen_excel.py TICKER XLS_FILENAME")
        sys.exit(1)

    ticker = sys.argv[1].upper()
    xls_filename = sys.argv[2]

    final_output = load_final_output(ticker)

    # Create or append to xls file
    writer = create_xls(xls_filename)

    # Now write data to each sheet
    write_summary_sheet(writer, final_output)
    write_company_description(writer, final_output)
    write_analyses_sheet(writer, final_output)
    write_balance_sheet_sheet(writer, final_output)
    write_profit_desc_sheet(writer, final_output)
    write_qualities_sheet(writer, final_output)
    write_industry_sheet(writer, final_output)
    write_segmentation_sheet(writer, final_output)
    write_hist_pricing_sheet(writer, final_output)
    write_valuation_sheet(writer, final_output, ticker)
    generate_config_note(ticker, writer.book)

    # Apply formatting: set font to Arial size 10 for non-formatted cells and remove gridlines
    format_workbook(writer)

    # Save changes
    writer.close()
    print(f"Data for {ticker} written to {xls_filename} successfully.")
